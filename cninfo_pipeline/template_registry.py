from __future__ import annotations

from dataclasses import dataclass
from functools import lru_cache
from pathlib import Path

from openpyxl import load_workbook

from .paths import resolve_project_root


TEMPLATE_GLOB_PATTERNS = (
    "cninfo_pipeline/*模版*.xlsx",
    "cninfo_pipeline/*模板*.xlsx",
    "cninfo_pipeline/templates/*.xlsx",
)
TEMPLATE_ID_ALIASES = {
    "工商银行财务报表模版": "银行财务报表模版",
    "长江电力年度报告财务报表模版": "公司财务报表模版",
    "bank": "银行财务报表模版",
    "company": "公司财务报表模版",
    "bank_template": "银行财务报表模版",
    "company_template": "公司财务报表模版",
}
DISPLAY_NAME_OVERRIDES = {
    "银行财务报表模版": "银行",
    "公司财务报表模版": "公司",
}


@dataclass(frozen=True)
class TemplateSpec:
    template_id: str
    display_name: str
    kind: str
    path: Path


@dataclass(frozen=True)
class BuiltinTemplate:
    template_id: str
    display_name: str
    kind: str
    relative_path: str


BUILTIN_TEMPLATES = (
    BuiltinTemplate(
        template_id="公司财务报表模版",
        display_name="公司",
        kind="company",
        relative_path="cninfo_pipeline/templates/company_template.xlsx",
    ),
    BuiltinTemplate(
        template_id="银行财务报表模版",
        display_name="银行",
        kind="bank",
        relative_path="cninfo_pipeline/templates/bank_template.xlsx",
    ),
)


def _normalize_marker(value: str) -> str:
    return "".join(str(value).replace("\xa0", "").split())


def _guess_template_kind(path: Path) -> str:
    workbook = load_workbook(path, read_only=True, data_only=False)
    try:
        markers: set[str] = set()
        for sheet in workbook.worksheets[:3]:
            for row in range(1, min(sheet.max_row, 25) + 1):
                markers.add(_normalize_marker(str(sheet.cell(row, 1).value or "")))
        joined_titles = " ".join(sheet.title for sheet in workbook.worksheets)
    finally:
        workbook.close()

    bank_markers = {
        "现金及存放中央银行款项",
        "吸收存款及同业存放",
        "利息净收入",
        "手续费及佣金净收入",
        "发行债务证券所收到的现金",
    }
    if "银行" in joined_titles or markers & bank_markers:
        return "bank"
    return "company"


def _build_display_name(path: Path, kind: str) -> str:
    if path.stem in DISPLAY_NAME_OVERRIDES:
        return DISPLAY_NAME_OVERRIDES[path.stem]
    prefix = "银行财务" if kind == "bank" else "公司财报"
    return f"{prefix} - {path.stem}"


@lru_cache(maxsize=1)
def discover_templates() -> tuple[TemplateSpec, ...]:
    project_root = resolve_project_root()
    templates: list[TemplateSpec] = []
    seen_paths: set[Path] = set()
    seen_template_ids: set[str] = set()

    for builtin in BUILTIN_TEMPLATES:
        path = (project_root / builtin.relative_path).resolve()
        if not path.exists():
            continue
        templates.append(
            TemplateSpec(
                template_id=builtin.template_id,
                display_name=builtin.display_name,
                kind=builtin.kind,
                path=path,
            )
        )
        seen_paths.add(path)
        seen_template_ids.add(builtin.template_id)

    for pattern in TEMPLATE_GLOB_PATTERNS:
        for path in sorted(project_root.glob(pattern)):
            resolved = path.resolve()
            if resolved in seen_paths:
                continue
            template_id = TEMPLATE_ID_ALIASES.get(path.stem, path.stem)
            if template_id in seen_template_ids:
                continue
            kind = _guess_template_kind(resolved)
            templates.append(
                TemplateSpec(
                    template_id=template_id,
                    display_name=_build_display_name(path, kind),
                    kind=kind,
                    path=resolved,
                )
            )
            seen_paths.add(resolved)
            seen_template_ids.add(template_id)

    return tuple(templates)


def template_data_files(project_root: Path | None = None) -> list[tuple[str, str]]:
    root = project_root or resolve_project_root()
    datas: list[tuple[str, str]] = []
    seen: set[Path] = set()

    for template in discover_templates():
        resolved = template.path.resolve()
        if resolved in seen:
            continue
        seen.add(resolved)
        destination = resolved.relative_to(root).parent.as_posix()
        datas.append((str(resolved), destination))

    return datas


def available_template_ids() -> tuple[str, ...]:
    template_ids = [template.template_id for template in discover_templates()]
    return tuple(template_ids + list(TEMPLATE_ID_ALIASES))


def default_template_id() -> str | None:
    templates = discover_templates()
    if not templates:
        return None
    for template in templates:
        if template.kind == "company":
            return template.template_id
    return templates[0].template_id


def resolve_template(template_id: str | None) -> TemplateSpec:
    templates = discover_templates()
    if not templates:
        raise FileNotFoundError("未找到可用的 Excel 模板，请把模板文件放到 cninfo_pipeline 目录。")

    requested = str(template_id or default_template_id() or "").strip()
    requested = TEMPLATE_ID_ALIASES.get(requested, requested)
    if requested:
        for template in templates:
            if template.template_id == requested:
                return template

    fallback_id = default_template_id()
    if fallback_id:
        for template in templates:
            if template.template_id == fallback_id:
                return template
    return templates[0]
