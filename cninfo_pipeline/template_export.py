from __future__ import annotations

from copy import copy
import re
import unicodedata
from dataclasses import dataclass
from datetime import date
from pathlib import Path
from typing import TYPE_CHECKING, Callable

from openpyxl import load_workbook
from openpyxl.styles import Alignment, Border, Font, PatternFill, Side
from openpyxl.utils import get_column_letter

from .client import CompanyRecord
from .constants import DEFAULT_UNIT_LABEL, UNIT_SCALE_MAP
from .official_source import official_label_keys
from .paths import ensure_writable_dir
from .service import DIRECT_FIELD_MAP, normalize_unit_label
from .template_registry import TemplateSpec, resolve_template

if TYPE_CHECKING:
    from .official_source import OfficialAnnualReportSource


Resolver = Callable[[dict], object | None]
DEFAULT_ANNUAL_PERIODS = 2
DATE_PATTERN = re.compile(r"\d{4}[-/]\d{1,2}[-/]\d{1,2}")
HEADER_FILL = PatternFill(fill_type="solid", fgColor="D9EAF7")
HEADER_FONT = Font(bold=True, color="1F2937")
DERIVED_FILL = PatternFill(fill_type="solid", fgColor="FFF2CC")
SECTION_FILL_BY_STATEMENT = {
    "balance": PatternFill(fill_type="solid", fgColor="EAF2F8"),
    "income": PatternFill(fill_type="solid", fgColor="EAF2F8"),
    "cash": PatternFill(fill_type="solid", fgColor="EAF2F8"),
}
SECTION_FONT = Font(bold=True, color="1F2937")
NOTE_HEADER = "注释"
NOTE_SUBHEADER = "来源"
MISSING_REASON_SHEET = "空白项说明"
SEPARATOR_SIDE = Side(style="medium", color="000000")
STATEMENT_SHEET_SUFFIX = {
    "balance": "资产负债表",
    "income": "利润表",
    "cash": "现金流量表",
}
SECTION_LABEL_EXACT = {
    "资产",
    "项目资产",
    "负债",
    "所有者权益",
    "股东权益",
    "流动资产",
    "非流动资产",
    "流动负债",
    "非流动负债",
    "经营活动产生的现金流量",
    "投资活动产生的现金流量",
    "筹资活动产生的现金流量",
    "现金及现金等价物净增加额",
    "每股收益",
    "以后将重分类进损益",
    "以后不能重分类进损益",
    "补充资料",
    "补充项目",
}
SECTION_LABEL_EXACT.update(
    {
        "项目",
        "项目资产",
        "金融投资",
        "金融投资：",
        "营业支出",
        "归属于",
        "归属于：",
        "其他综合收益",
        "股东权益",
    }
)


@dataclass(frozen=True)
class SupplementalItem:
    label: str
    resolver: Resolver


@dataclass(frozen=True)
class FieldCatalogItem:
    field_name: str
    label: str
    template_kind: str
    statement_type: str


def _resolver_with_metadata(
    resolver: Resolver,
    *,
    kind: str,
    source_fields: tuple[str, ...] = (),
    source_labels: tuple[str, ...] = (),
) -> Resolver:
    setattr(resolver, "_resolver_kind", kind)
    setattr(resolver, "_source_fields", tuple(dict.fromkeys(source_fields)))
    setattr(resolver, "_source_labels", tuple(dict.fromkeys(source_labels)))
    return resolver


def resolver_kind(resolver: Resolver | None) -> str:
    return str(getattr(resolver, "_resolver_kind", "custom"))


def resolver_source_fields(resolver: Resolver | None) -> tuple[str, ...]:
    return tuple(getattr(resolver, "_source_fields", ()))


def resolver_source_labels(resolver: Resolver | None) -> tuple[str, ...]:
    return tuple(getattr(resolver, "_source_labels", ()))


def is_derived_resolver(resolver: Resolver | None) -> bool:
    return resolver_kind(resolver) in {"sum", "subtract", "sum_available", "official_sum"}


def field(field_name: str) -> Resolver:
    return _resolver_with_metadata(
        lambda record: record.get(field_name),
        kind="field",
        source_fields=(field_name,),
    )


def sum_fields(*field_names: str) -> Resolver:
    def resolver(record: dict) -> object | None:
        values = [record.get(field_name) for field_name in field_names]
        numeric = [value for value in values if isinstance(value, (int, float))]
        return sum(numeric) if numeric else None

    return _resolver_with_metadata(
        resolver,
        kind="sum",
        source_fields=tuple(field_names),
    )


def subtract_fields(total_field: str, *part_fields: str) -> Resolver:
    def resolver(record: dict) -> object | None:
        total_value = record.get(total_field)
        if not isinstance(total_value, (int, float)):
            return None
        parts = [record.get(field_name) for field_name in part_fields]
        numeric_parts = [value for value in parts if isinstance(value, (int, float))]
        return total_value - sum(numeric_parts)

    return _resolver_with_metadata(
        resolver,
        kind="subtract",
        source_fields=(total_field, *part_fields),
    )


def first_available(*resolvers: Resolver) -> Resolver:
    def resolver(record: dict) -> object | None:
        for candidate in resolvers:
            value = candidate(record)
            if value is not None:
                return value
        return None

    source_fields: list[str] = []
    for candidate in resolvers:
        source_fields.extend(resolver_source_fields(candidate))
    return _resolver_with_metadata(
        resolver,
        kind="first_available",
        source_fields=tuple(source_fields),
    )


def sum_available_fields(*field_names: str) -> Resolver:
    def resolver(record: dict) -> object | None:
        values = [record.get(field_name) for field_name in field_names]
        numeric = [value for value in values if isinstance(value, (int, float))]
        return sum(numeric) if numeric else None

    return _resolver_with_metadata(
        resolver,
        kind="sum_available",
        source_fields=tuple(field_names),
    )


def placeholder(kind: str, *source_fields: str) -> Resolver:
    return _resolver_with_metadata(
        lambda _record: None,
        kind=kind,
        source_fields=tuple(source_fields),
    )


def aggregate_only(*source_fields: str) -> Resolver:
    return placeholder("aggregate_only", *source_fields)


def formula_required(*source_fields: str) -> Resolver:
    return placeholder("formula_required", *source_fields)


def official_value(*labels: str) -> Resolver:
    canonical_labels = tuple(canonical_label(label) for label in labels)

    def resolver(record: dict) -> object | None:
        official_rows = record.get("__official_rows__", {})
        for label_key in canonical_labels:
            if label_key in official_rows:
                return official_rows[label_key]
        return None

    return _resolver_with_metadata(
        resolver,
        kind="official",
        source_labels=tuple(labels),
    )


def official_sum(*labels: str) -> Resolver:
    canonical_labels = tuple(canonical_label(label) for label in labels)

    def resolver(record: dict) -> object | None:
        official_rows = record.get("__official_rows__", {})
        values = [official_rows.get(label_key) for label_key in canonical_labels]
        numeric = [value for value in values if isinstance(value, (int, float))]
        return sum(numeric) if numeric else None

    return _resolver_with_metadata(
        resolver,
        kind="official_sum",
        source_labels=tuple(labels),
    )
def normalize_label(value: object | None) -> str:
    text = str(value or "").strip()
    replacements = {
        "\xa0": "",
        " ": "",
        "　": "",
        "（": "",
        "）": "",
        "(": "",
        ")": "",
        "：": "",
        ":": "",
        "、": "",
        "，": "",
        ",": "",
        "。": "",
        "“": "",
        "”": "",
        "‘": "",
        "’": "",
        "－": "",
        "-": "",
        "/": "",
        "／": "",
        "·": "",
        "\n": "",
        "\r": "",
        "\t": "",
    }
    for old, new in replacements.items():
        text = text.replace(old, new)
    return text


LABEL_FIXUPS = {
    normalize_label("长期待滩费用"): normalize_label("长期待摊费用"),
    normalize_label("其中：营业收入"): normalize_label("营业收入"),
    normalize_label("其中：营业成本"): normalize_label("营业成本"),
    normalize_label("其中：应收利息"): normalize_label("应收利息"),
    normalize_label("其中：应付利息"): normalize_label("应付利息"),
    normalize_label("实收资本（或股本）"): normalize_label("实收资本(或股本)"),
    normalize_label("所有者权益（或股东权益）"): normalize_label("所有者权益"),
    normalize_label("归属于母公司所有者权益 （或股东权益）合计"): normalize_label("归属于母公司所有者权益（或股东权益）合计"),
    normalize_label("所有者权益（或股东权益）合计"): normalize_label("所有者权益(或股东权益)合计"),
    normalize_label("负债和所有者权益（或股东权益）总计"): normalize_label("负债和所有者权益(或股东权益)总计"),
    normalize_label("经营活动产生的现金量"): normalize_label("经营活动产生的现金流量"),
    normalize_label("经营活动产生的现金量："): normalize_label("经营活动产生的现金流量："),
    normalize_label("客户存款和同业存放款项净增加額"): normalize_label("客户存款和同业存放款项净增加额"),
    normalize_label("向其他金融机构拆入资金净增加額"): normalize_label("向其他金融机构拆入资金净增加额"),
    normalize_label("收取利息、手续费及佣金的现"): normalize_label("收取利息、手续费及佣金的现金"),
    normalize_label("取得惜款收到的现金"): normalize_label("取得借款收到的现金"),
    normalize_label("收到其他与等资活动有关的现金"): normalize_label("收到其他与筹资活动有关的现金"),
    normalize_label("支付其他与等资活动有关的现金"): normalize_label("支付其他与筹资活动有关的现金"),
    normalize_label("投资动产生的现金流量净额"): normalize_label("投资活动产生的现金流量净额"),
    normalize_label("五、现金及现金等价物净增加额加期初现金及现金等价物余额"): normalize_label("现金及现金等价物净增加额"),
    normalize_label("六、期末现金及现金等价物余额"): normalize_label("期末现金及现金等价物余额"),
    normalize_label("五、现金及现金等价物净变动额加：年初现金及现金等价物余额"): normalize_label("现金及现金等价物净变动额"),
    normalize_label("六、年末现金及现金等价物余额"): normalize_label("年末现金及现金等价物余额"),
    normalize_label("四、汇率变动对现金及现金等价物"): normalize_label("四、汇率变动对现金及现金等价物的影响"),
}


def canonical_label(value: object | None) -> str:
    normalized = normalize_label(value)
    return LABEL_FIXUPS.get(normalized, normalized)


NON_SCALED_LABELS = {
    canonical_label("基本每股收益（元/股）"),
    canonical_label("稀释每股收益（元/股）"),
    canonical_label("基本每股收益（人民币元）"),
    canonical_label("稀释每股收益（人民币元）"),
}

OFFICIAL_OVERRIDE_GUARDED_LABELS = {
    canonical_label("所有者权益(或股东权益)合计"),
}


COMPANY_BALANCE_RESOLVERS: dict[str, Resolver] = {
    canonical_label("货币资金"): field("F006N"),
    canonical_label("交易性金融资产"): field("F117N"),
    canonical_label("衍生金融资产"): field("F080N"),
    canonical_label("应收票据"): field("F008N"),
    canonical_label("应收账款"): field("F009N"),
    canonical_label("应收款项融资"): field("F110N"),
    canonical_label("预付款项"): field("F010N"),
    canonical_label("其他应收款"): field("F011N"),
    canonical_label("应收利息"): field("F012N"),
    canonical_label("应收股利"): field("F013N"),
    canonical_label("买入返售金融资产"): field("F016N"),
    canonical_label("存货"): field("F015N"),
    canonical_label("合同资产"): field("F119N"),
    canonical_label("一年内到期的非流动资产"): field("F017N"),
    canonical_label("其他流动资产"): field("F018N"),
    canonical_label("流动资产合计"): field("F019N"),
    canonical_label("发放贷款和垫款"): field("F113N"),
    canonical_label("长期应收款"): field("F022N"),
    canonical_label("长期股权投资"): field("F023N"),
    canonical_label("其他权益工具投资"): field("F111N"),
    canonical_label("其他非流动金融资产"): field("F112N"),
    canonical_label("投资性房地产"): field("F024N"),
    canonical_label("固定资产"): field("F025N"),
    canonical_label("在建工程"): field("F026N"),
    canonical_label("使用权资产"): field("F121N"),
    canonical_label("无形资产"): field("F031N"),
    canonical_label("开发支出"): field("F032N"),
    canonical_label("商誉"): field("F033N"),
    canonical_label("长期待摊费用"): field("F034N"),
    canonical_label("递延所得税资产"): field("F035N"),
    canonical_label("其他非流动资产"): field("F036N"),
    canonical_label("非流动资产合计"): field("F037N"),
    canonical_label("资产总计"): field("F038N"),
    canonical_label("短期借款"): field("F039N"),
    canonical_label("交易性金融负债"): field("F090N"),
    canonical_label("衍生金融负债"): field("F091N"),
    canonical_label("应付票据"): field("F041N"),
    canonical_label("应付账款"): field("F042N"),
    canonical_label("预收款项"): field("F043N"),
    canonical_label("合同负债"): field("F115N"),
    canonical_label("应付职工薪酬"): field("F044N"),
    canonical_label("应交税费"): field("F045N"),
    canonical_label("其他应付款"): subtract_fields("F048N", "F046N", "F047N"),
    canonical_label("应付利息"): field("F046N"),
    canonical_label("应付股利"): field("F047N"),
    canonical_label("一年内到期的非流动负债"): field("F050N"),
    canonical_label("其他流动负债"): field("F051N"),
    canonical_label("流动负债合计"): field("F052N"),
    canonical_label("长期借款"): field("F053N"),
    canonical_label("应付债券"): field("F054N"),
    canonical_label("租赁负债"): field("F122N"),
    canonical_label("长期应付款"): field("F076N"),
    canonical_label("长期应付职工薪酬"): field("F055N"),
    canonical_label("预计负债"): field("F057N"),
    canonical_label("递延收益"): field("F075N"),
    canonical_label("递延所得税负债"): field("F058N"),
    canonical_label("其他非流动负债"): field("F059N"),
    canonical_label("非流动负债合计"): field("F060N"),
    canonical_label("负债合计"): field("F061N"),
    canonical_label("实收资本(或股本)"): field("F062N"),
    canonical_label("资本公积"): field("F063N"),
    canonical_label("减库存股"): field("F066N"),
    canonical_label("其他综合收益"): field("F074N"),
    canonical_label("专项储备"): field("F068N"),
    canonical_label("盈余公积"): field("F064N"),
    canonical_label("一般风险准备"): field("F069N"),
    canonical_label("未分配利润"): field("F065N"),
    canonical_label("归属于母公司所有者权益（或股东权益）合计"): field("F073N"),
    canonical_label("少数股东权益"): field("F067N"),
    canonical_label("所有者权益(或股东权益)合计"): field("F070N"),
    canonical_label("负债和所有者权益(或股东权益)总计"): field("F071N"),
}


COMPANY_INCOME_RESOLVERS: dict[str, Resolver] = {
    canonical_label("一、营业总收入"): first_available(field("F035N"), field("F006N")),
    canonical_label("营业收入"): field("F006N"),
    canonical_label("二、营业总成本"): field("F036N"),
    canonical_label("营业成本"): field("F007N"),
    canonical_label("税金及附加"): field("F008N"),
    canonical_label("销售费用"): field("F009N"),
    canonical_label("管理费用"): field("F010N"),
    canonical_label("研发费用"): field("F056N"),
    canonical_label("财务费用"): field("F012N"),
    canonical_label("加其他收益"): field("F051N"),
    canonical_label("投资收益损失以号填列"): field("F015N"),
    canonical_label("其中对联营企业和合营企业的投资收益"): field("F016N"),
    canonical_label("汇兑收益损失以填列"): field("F023N"),
    canonical_label("公允价值变动收益损失以号填列"): field("F014N"),
    canonical_label("信用减值损失损失以号填列"): field("F064N"),
    canonical_label("资产减值损失损失以号填列"): field("F065N"),
    canonical_label("资产处置收益损失以号填列"): field("F059N"),
    canonical_label("三、营业利润亏损以号填列"): field("F018N"),
    canonical_label("加营业外收入"): field("F020N"),
    canonical_label("减营业外支出"): field("F021N"),
    canonical_label("四、利润总额亏损以号填列"): field("F024N"),
    canonical_label("减所得税费用"): field("F025N"),
    canonical_label("五、净利润净亏损以号填列"): field("F027N"),
    canonical_label("1持续经营冲利润净亏损以号填列"): first_available(field("F060N"), field("F027N")),
    canonical_label("1归属于母公司股东的净利润净亏损以号填列"): field("F028N"),
    canonical_label("2少数股东损益净亏损以号填列"): field("F029N"),
    canonical_label("六、其他综合收益的税后冲额"): field("F038N"),
    canonical_label("一归属母公司所有者的其他综合收益的税后净额"): field("F066N"),
    canonical_label("二归属于少数股东的其他综合收益的税后净额"): field("F067N"),
    canonical_label("七、综合收益总额"): field("F039N"),
    canonical_label("一归属于母公司所有者的综合收益总额"): field("F040N"),
    canonical_label("二归属于少数股东的综合收益总额"): field("F041N"),
    canonical_label("一基本每股收益元股"): field("F031N"),
    canonical_label("二稀释每股收益元股"): field("F032N"),
}


COMPANY_CASH_RESOLVERS: dict[str, Resolver] = {
    canonical_label("销售商品提供劳务收到的现金"): field("F006N"),
    canonical_label("收到的税费返还"): field("F007N"),
    canonical_label("收到其他与经营活动有关的现金"): field("F008N"),
    canonical_label("经营活动现金流入小计"): field("F009N"),
    canonical_label("购买商品接受劳务支付的现金"): field("F010N"),
    canonical_label("支付给职工及为职工支付的现金"): field("F011N"),
    canonical_label("支付的各项税费"): field("F012N"),
    canonical_label("支付其他与经营活动有关的现金"): field("F013N"),
    canonical_label("经营活动现金流出小计"): field("F014N"),
    canonical_label("经营活动产生的现金流量净额"): field("F015N"),
    canonical_label("收回投资收到的现金"): field("F016N"),
    canonical_label("取得投资收益收到的现金"): field("F017N"),
    canonical_label("处置固定资产无形资产和其他长期资产收回的现金净额"): field("F018N"),
    canonical_label("处置子公司及其他营业单位收到的现金净额"): field("F019N"),
    canonical_label("收到其他与投资活动有关的现金"): field("F020N"),
    canonical_label("投资活动现金流入小计"): field("F021N"),
    canonical_label("购建固定资产无形资产和其他长期资产支付的现金"): field("F022N"),
    canonical_label("投资支付的现金"): field("F023N"),
    canonical_label("取得子公司及其他营业单位支付的现金净额"): field("F024N"),
    canonical_label("支付其他与投资活动有关的现金"): field("F025N"),
    canonical_label("投资活动现金流出小计"): field("F026N"),
    canonical_label("投资活动产生的现金流量净额"): field("F027N"),
    canonical_label("吸收投资收到的现金"): field("F028N"),
    canonical_label("其中子公司吸收少数股东投资收到的现金"): field("F089N"),
    canonical_label("取得借款收到的现金"): field("F029N"),
    canonical_label("收到其他与筹资活动有关的现金"): field("F030N"),
    canonical_label("筹资活动现金流入小计"): field("F031N"),
    canonical_label("偿还债务支付的现金"): field("F032N"),
    canonical_label("分配股利利润或偿付利息支付的现金"): field("F033N"),
    canonical_label("其中子公司支付给少数股东的股利利润"): field("F091N"),
    canonical_label("支付其他与筹资活动有关的现金"): field("F034N"),
    canonical_label("筹资活动现金流出小计"): field("F035N"),
    canonical_label("筹资活动产生的现金流量净额"): field("F036N"),
    canonical_label("四、汇率变动对现金及现金等价物的影响"): field("F037N"),
    canonical_label("现金及现金等价物净增加额"): field("F039N"),
    canonical_label("期初现金及现金等价物余额"): field("F040N"),
    canonical_label("期末现金及现金等价物余额"): field("F041N"),
}


BANK_BALANCE_RESOLVERS: dict[str, Resolver] = {
    canonical_label("现金及存放中央银行款项"): field("F006N"),
    canonical_label("衍生金融资产"): field("F035N"),
    canonical_label("买入返售款项"): field("F117N"),
    canonical_label("长期股权投资"): field("F023N"),
    canonical_label("固定资产"): field("F025N"),
    canonical_label("在建工程"): field("F026N"),
    canonical_label("递延所得税资产"): field("F087N"),
    canonical_label("资产总计"): field("F038N"),
    canonical_label("拆入资金"): field("F089N"),
    canonical_label("以公允价值计量且其变动计入当期损益的金融负债"): first_available(field("F040N"), field("F113N")),
    canonical_label("衍生金融负债"): field("F090N"),
    canonical_label("卖出回购款项"): field("F091N"),
    canonical_label("应付职工薪酬"): field("F044N"),
    canonical_label("应交税费"): field("F045N"),
    canonical_label("已发行债务证券"): field("F054N"),
    canonical_label("递延所得税负债"): field("F058N"),
    canonical_label("负债合计"): field("F061N"),
    canonical_label("股本"): field("F062N"),
    canonical_label("其他权益工具"): field("F103N"),
    canonical_label("优先股"): field("F104N"),
    canonical_label("永续债"): field("F105N"),
    canonical_label("资本公积"): field("F063N"),
    canonical_label("其他综合收益"): field("F074N"),
    canonical_label("盈余公积"): field("F064N"),
    canonical_label("一般准备"): field("F076N"),
    canonical_label("未分配利润"): field("F065N"),
    canonical_label("归属于母公司股东的权益"): field("F073N"),
    canonical_label("少数股东权益"): field("F067N"),
    canonical_label("股东权益合计"): field("F070N"),
    canonical_label("负债及股东权益总计"): field("F071N"),
}


BANK_INCOME_RESOLVERS: dict[str, Resolver] = {
    canonical_label("利息净收入"): field("F033N"),
    canonical_label("手续费及佣金净收入"): field("F042N"),
    canonical_label("投资收益"): field("F015N"),
    canonical_label("其中对联营及合营企业的投资收益"): field("F016N"),
    canonical_label("公允价值变动净收益损失"): field("F014N"),
    canonical_label("汇兑及汇率产品净损失"): field("F017N"),
    canonical_label("其他业务收入"): field("F013N"),
    canonical_label("营业收入"): field("F035N"),
    canonical_label("税金及附加"): field("F008N"),
    canonical_label("业务及管理费"): field("F057N"),
    canonical_label("资产减值损失"): field("F010N"),
    canonical_label("营业支出"): field("F036N"),
    canonical_label("营业利润"): field("F018N"),
    canonical_label("加营业外收入"): field("F020N"),
    canonical_label("减营业外支出"): field("F021N"),
    canonical_label("税前利润"): field("F024N"),
    canonical_label("减所得税费用"): field("F025N"),
    canonical_label("净利润"): field("F027N"),
    canonical_label("母公司股东"): field("F028N"),
    canonical_label("少数股东"): field("F029N"),
    canonical_label("本年净利润"): field("F027N"),
    canonical_label("其他综合收益的税后净额"): field("F038N"),
    canonical_label("本年其他综合收益小计"): field("F038N"),
    canonical_label("本年综合收益总额"): field("F039N"),
    canonical_label("基本每股收益人民币元"): field("F031N"),
    canonical_label("稀释每股收益人民币元"): field("F032N"),
}


BANK_CASH_RESOLVERS: dict[str, Resolver] = {
    canonical_label("收取的利息手续费及佣金的现金"): field("F081N"),
    canonical_label("经营活动现金流入小计"): field("F009N"),
    canonical_label("经营活动现金流出小计"): field("F014N"),
    canonical_label("经营活动产生的现金流量净额"): first_available(field("F015N"), field("F060N")),
    canonical_label("客户贷款及垫款净额"): field("F084N"),
    canonical_label("支付的利息手续费及佣金的现金"): field("F087N"),
    canonical_label("支付给职工以及为职工支付的现金支付的各项税费"): sum_fields("F011N", "F012N"),
    canonical_label("收回投资收到的现金"): field("F016N"),
    canonical_label("取得投资收益收到的现金"): field("F017N"),
    canonical_label("处置固定资产无形资产和其他长期资产不含抵债资产收回的现金"): field("F018N"),
    canonical_label("收到其他与投资活动有关的现金"): field("F020N"),
    canonical_label("投资活动现金流入小计"): field("F021N"),
    canonical_label("投资支付的现金"): field("F023N"),
    canonical_label("购建固定资产无形资产和其他长期资产支付的现金"): field("F022N"),
    canonical_label("投资活动现金流出小计"): field("F026N"),
    canonical_label("投资活动产生的现金流量净额"): field("F027N"),
    canonical_label("发行债务证券所收到的现金"): field("F076N"),
    canonical_label("筹资活动现金流入小计"): field("F031N"),
    canonical_label("筹资活动产生的现金流量净额"): field("F036N"),
    canonical_label("偿还债务证券所支付的现金"): field("F032N"),
    canonical_label("支付其他与筹资活动有关的现金"): field("F034N"),
    canonical_label("筹资活动现金流出小计"): field("F035N"),
    canonical_label("四汇率变动对现金及现金等价物的影响"): field("F037N"),
    canonical_label("现金及现金等价物净变动额"): field("F071N"),
    canonical_label("年初现金及现金等价物余额"): field("F040N"),
    canonical_label("年末现金及现金等价物余额"): field("F041N"),
}

COMPANY_BALANCE_RESOLVERS.update(
    {
        canonical_label("应收票据及应收账款"): sum_fields("F008N", "F009N"),
        canonical_label("划分为持有待售的资产"): field("F118N"),
        canonical_label("债权投资"): aggregate_only("F022N"),
        canonical_label("其他债权投资"): aggregate_only("F022N"),
        canonical_label("可供出售金融资产"): aggregate_only("F111N"),
        canonical_label("持有至到期投资"): aggregate_only("F112N"),
        canonical_label("发放贷款及垫款"): field("F113N"),
        canonical_label("固定资产及清理"): sum_available_fields("F025N", "F028N"),
        canonical_label("工程物资"): field("F027N"),
        canonical_label("固定资产净额"): field("F025N"),
        canonical_label("固定资产清理"): field("F028N"),
        canonical_label("应付票据及应付账款"): sum_fields("F041N", "F042N"),
        canonical_label("吸收存款同业存放"): placeholder("no_api_field"),
        canonical_label("应付手续费及佣金"): aggregate_only("F115N"),
        canonical_label("一年内的递延收益"): field("F116N"),
        canonical_label("应付短期债券"): field("F114N"),
        canonical_label("预计非流动负债"): field("F057N"),
        canonical_label("专项应付款"): field("F077N"),
        canonical_label("长期递延收益"): field("F075N"),
        canonical_label("归属于母公司股东权益合计"): field("F073N"),
    }
)

COMPANY_INCOME_RESOLVERS.update(
    {
        canonical_label("一、营业收入"): field("F006N"),
        canonical_label("营业税金及附加"): field("F008N"),
        canonical_label("公允价值变动收益"): field("F014N"),
        canonical_label("投资收益"): field("F015N"),
        canonical_label("汇兑收益"): field("F023N"),
        canonical_label("其中：利息费用"): official_value("利息费用"),
        canonical_label("利息收入"): official_value("利息收入"),
        canonical_label("加：其他收益"): field("F051N"),
        canonical_label("信用减值损失"): field("F064N"),
        canonical_label("资产处置收益"): field("F059N"),
        canonical_label("三、营业利润"): field("F018N"),
        canonical_label("加:营业外收入"): field("F020N"),
        canonical_label("减:营业外支出"): field("F021N"),
        canonical_label("减：营业外支出"): field("F021N"),
        canonical_label("其中：非流动资产处置损失"): formula_required("F021N", "F059N"),
        canonical_label("四、利润总额"): field("F024N"),
        canonical_label("减：所得税费用"): field("F025N"),
        canonical_label("五、净利润"): field("F027N"),
        canonical_label("归属于母公司所有者的净利润"): field("F028N"),
        canonical_label("少数股东损益"): field("F029N"),
        canonical_label("基本每股收益元股"): field("F031N"),
        canonical_label("稀释每股收益元股"): field("F032N"),
        canonical_label("基本每股收益元股"): field("F031N"),
        canonical_label("七、其他综合收益"): field("F038N"),
        canonical_label("八、综合收益总额"): field("F039N"),
        canonical_label("归属于母公司所有者的综合收益总额"): field("F040N"),
        canonical_label("归属于少数股东的综合收益总额"): field("F041N"),
        canonical_label("资产减值损失"): field("F065N"),
    }
)

COMPANY_CASH_RESOLVERS.update(
    {
        canonical_label("收到的其他与经营活动有关的现金"): field("F008N"),
        canonical_label("支付给职工以及为职工支付的现金"): field("F011N"),
        canonical_label("支付的其他与经营活动有关的现金"): field("F013N"),
        canonical_label("收回投资所收到的现金"): field("F016N"),
        canonical_label("取得投资收益所收到的现金"): field("F017N"),
        canonical_label("处置固定资产、无形资产和其他长期资产所收回的现金净额"): field("F018N"),
        canonical_label("处置子公司及其他营业单位收到的现金净额"): field("F019N"),
        canonical_label("收到的其他与投资活动有关的现金"): field("F020N"),
        canonical_label("购建固定资产、无形资产和其他长期资产所支付的现金"): field("F022N"),
        canonical_label("投资所支付的现金"): field("F023N"),
        canonical_label("取得子公司及其他营业单位支付的现金净额"): field("F024N"),
        canonical_label("支付的其他与投资活动有关的现金"): field("F025N"),
        canonical_label("吸收投资收到的现金"): field("F028N"),
        canonical_label("其中：子公司吸收少数股东投资收到的现金"): field("F089N"),
        canonical_label("发行债券收到的现金"): aggregate_only("F029N"),
        canonical_label("收到其他与筹资活动有关的现金"): field("F030N"),
        canonical_label("分配股利、利润或偿付利息所支付的现金"): field("F033N"),
        canonical_label("其中：子公司支付给少数股东的股利、利润"): field("F091N"),
        canonical_label("四、汇率变动对现金及现金等价物的影响"): field("F037N"),
        canonical_label("加:期初现金及现金等价物余额"): field("F040N"),
        canonical_label("六、期末现金及现金等价物余额"): field("F041N"),
        canonical_label("少数股东权益"): field("F029N"),
    }
)

BANK_BALANCE_RESOLVERS.update(
    {
        canonical_label("现金及存放中央银行款项"): official_value("现金及存放中央银行款项"),
        canonical_label("发放贷款及垫款"): official_value("发放贷款及垫款"),
        canonical_label("买入返售金融资产"): official_value("买入返售金融资产"),
        canonical_label("同业存入及拆入"): official_sum("其中:同业存放款项", "拆入资金"),
        canonical_label("其中:同业存放款项"): official_value("其中:同业存放款项"),
        canonical_label("客户存款(吸收存款)"): official_value("客户存款(吸收存款)"),
        canonical_label("拆入资金"): official_value("拆入资金"),
        canonical_label("衍生金融工具资产"): field("F035N"),
        canonical_label("固定资产合计"): field("F025N"),
        canonical_label("无形资产"): field("F031N"),
        canonical_label("商誉"): field("F033N"),
        canonical_label("递延税款借项"): field("F087N"),
        canonical_label("投资性房地产"): field("F024N"),
        canonical_label("衍生金融工具负债"): field("F090N"),
        canonical_label("交易性金融负债"): first_available(field("F040N"), field("F113N")),
        canonical_label("卖出回购金融资产款"): field("F091N"),
        canonical_label("应付职工薪酬"): field("F044N"),
        canonical_label("应交税费"): field("F045N"),
        canonical_label("应付利息"): field("F046N"),
        canonical_label("应付账款"): field("F042N"),
        canonical_label("应付债券"): field("F054N"),
        canonical_label("递延所得税负债"): field("F058N"),
        canonical_label("预计负债"): field("F057N"),
        canonical_label("减:库藏股"): field("F066N"),
        canonical_label("外币报表折算差额"): aggregate_only("F074N"),
    }
)

BANK_INCOME_RESOLVERS.update(
    {
        canonical_label("一、营业收入"): field("F035N"),
        canonical_label("其中：利息收入"): aggregate_only("F033N"),
        canonical_label("减：利息支出"): aggregate_only("F033N"),
        canonical_label("其中:手续费及佣金收入"): aggregate_only("F042N"),
        canonical_label("减：手续费及佣金支出"): aggregate_only("F042N"),
        canonical_label("汇兑收益"): field("F017N"),
        canonical_label("投资净收益"): field("F015N"),
        canonical_label("其中:对联营公司的投资收益"): field("F016N"),
        canonical_label("公允价值变动净收益"): field("F014N"),
        canonical_label("营业税金及附加"): field("F008N"),
        canonical_label("研发费用"): placeholder("no_api_field"),
        canonical_label("其他业务支出"): placeholder("no_api_field"),
        canonical_label("三、营业利润"): field("F018N"),
        canonical_label("加:营业外收入"): field("F020N"),
        canonical_label("减:营业外支出"): field("F021N"),
        canonical_label("四、利润总额"): field("F024N"),
        canonical_label("减:所得税"): field("F025N"),
        canonical_label("五、净利润"): field("F027N"),
        canonical_label("归属于母公司的净利润"): field("F028N"),
        canonical_label("少数股东权益"): field("F029N"),
        canonical_label("基本每股收益元股"): field("F031N"),
        canonical_label("稀释每股收益元股"): field("F032N"),
        canonical_label("七、其他综合收益"): field("F038N"),
        canonical_label("八、综合收益总额"): field("F039N"),
        canonical_label("归属于母公司所有者的综合收益总额"): field("F040N"),
        canonical_label("归属于少数股东的综合收益总额"): field("F041N"),
    }
)

BANK_CASH_RESOLVERS.update(
    {
        canonical_label("收到其他与经营活动有关的现金"): field("F008N"),
        canonical_label("收取利息、手续费及佣金的现金"): field("F081N"),
        canonical_label("客户贷款及垫款净增加额"): field("F084N"),
        canonical_label("支付其他与经营活动有关的现金"): field("F013N"),
        canonical_label("支付利息、手续费及佣金的现金"): field("F087N"),
        canonical_label("客户存款和同业存放款项净增加额"): field("F072N"),
        canonical_label("向中央银行借款净增加额"): field("F073N"),
        canonical_label("向其他金融机构拆入资金净增加额"): field("F074N"),
        canonical_label("存放中央银行和同业款项净增加额"): field("F085N"),
        canonical_label("支付给职工以及为职工支付的现金"): field("F011N"),
        canonical_label("支付的各项税费"): field("F012N"),
        canonical_label("处置固定资产、无形资产及其他资产而收到的现金"): field("F018N"),
        canonical_label("取得子公司及其他营业单位所收到的现金净额"): field("F019N"),
        canonical_label("支付的其他与投资活动有关的现金"): field("F025N"),
        canonical_label("吸收投资所收到的现金"): field("F028N"),
        canonical_label("发行债券收到的现金"): field("F076N"),
        canonical_label("收到其他与筹资活动有关的现金"): field("F030N"),
        canonical_label("分配股利、利润或偿付利息支付的现金"): field("F033N"),
        canonical_label("其中:偿付利息所支付的现金"): aggregate_only("F033N"),
        canonical_label("五、现金及现金等价物净增加额"): field("F071N"),
        canonical_label("加:期初现金及现金等价物余额"): field("F040N"),
        canonical_label("六、期末现金及现金等价物余额"): field("F041N"),
        canonical_label("加:少数股东收益"): field("F029N"),
        canonical_label("经营活动现金流量净额"): first_available(field("F015N"), field("F060N")),
        canonical_label("现金的期末余额"): field("F041N"),
        canonical_label("减:现金的期初余额"): field("F040N"),
        canonical_label("现金等价物的期末余额"): field("F041N"),
        canonical_label("减：现金等价物的期初余额"): field("F040N"),
        canonical_label("现金及现金等价物净增加额"): field("F071N"),
    }
)

BANK_BALANCE_RESOLVERS.update(
    {
        canonical_label("现金"): official_value("现金"),
        canonical_label("贵金属"): official_value("贵金属"),
        canonical_label("存放中央银行款项"): official_value("存放中央银行款项"),
        canonical_label("存放同业和其他金融机构款项"): official_value("存放同业和其他金融机构款项"),
        canonical_label("拆出资金"): official_value("拆出资金"),
        canonical_label("买入返售金融资产"): field("F117N"),
        canonical_label("贷款和垫款"): official_value("贷款和垫款"),
        canonical_label("衍生金融资产"): field("F035N"),
        canonical_label("以公允价值计量且其变动计入当期损益的金融投资"): official_value(
            "以公允价值计量且其变动计入当期损益的金融投资"
        ),
        canonical_label("以摊余成本计量的债务工具投资"): official_value("以摊余成本计量的债务工具投资"),
        canonical_label("以公允价值计量且其变动计入其他综合收益的债务工具投资"): official_value(
            "以公允价值计量且其变动计入其他综合收益的债务工具投资"
        ),
        canonical_label("指定为以公允价值计量且其变动计入其他综合收益的权益工具投资"): official_value(
            "指定为以公允价值计量且其变动计入其他综合收益的权益工具投资"
        ),
        canonical_label("长期股权投资"): field("F023N"),
        canonical_label("投资性房地产"): official_value("投资性房地产"),
        canonical_label("固定资产"): field("F025N"),
        canonical_label("在建工程"): official_value("在建工程"),
        canonical_label("使用权资产"): official_value("使用权资产"),
        canonical_label("无形资产"): official_value("无形资产"),
        canonical_label("商誉"): official_value("商誉"),
        canonical_label("递延所得税资产"): field("F087N"),
        canonical_label("其他资产"): official_value("其他资产"),
        canonical_label("资产合计"): field("F038N"),
        canonical_label("向中央银行借款"): official_value("向中央银行借款"),
        canonical_label("同业和其他金融机构存放款项"): official_value("同业和其他金融机构存放款项"),
        canonical_label("卖出回购金融资产款"): field("F091N"),
        canonical_label("客户存款"): official_value("客户存款"),
        canonical_label("合同负债"): official_value("合同负债"),
        canonical_label("租赁负债"): official_value("租赁负债"),
        canonical_label("预计负债"): official_value("预计负债"),
        canonical_label("其他负债"): official_value("其他负债"),
        canonical_label("一般风险准备"): field("F076N"),
        canonical_label("其中：建议分配利润"): official_value("其中：建议分配利润"),
        canonical_label("归属于本行股东权益合计"): field("F073N"),
        canonical_label("其中：普通股少数股东权益"): official_value("其中：普通股少数股东权益"),
        canonical_label("股东权益合计"): field("F070N"),
        canonical_label("负债及股东权益总计"): field("F071N"),
    }
)

BANK_INCOME_RESOLVERS.update(
    {
        canonical_label("利息收入"): official_value("利息收入"),
        canonical_label("利息支出"): official_value("利息支出"),
        canonical_label("净利息收入"): field("F033N"),
        canonical_label("手续费及佣金收入"): official_value("手续费及佣金收入"),
        canonical_label("手续费及佣金支出"): official_value("手续费及佣金支出"),
        canonical_label("净手续费及佣金收入"): field("F042N"),
        canonical_label("投资收益"): field("F015N"),
        canonical_label("其中：对合营企业及联营企业的投资收益"): field("F016N"),
        canonical_label("以摊余成本计量的金融资产终止确认产生的收益"): official_value(
            "以摊余成本计量的金融资产终止确认产生的收益"
        ),
        canonical_label("公允价值变动收益"): field("F014N"),
        canonical_label("汇兑净收益"): field("F017N"),
        canonical_label("其他净收入"): official_value("其他净收入"),
        canonical_label("营业收入小计"): field("F035N"),
        canonical_label("税金及附加"): field("F008N"),
        canonical_label("业务及管理费"): field("F057N"),
        canonical_label("信用减值损失"): official_value("信用减值损失"),
        canonical_label("其他资产减值损失"): official_value("其他资产减值损失"),
        canonical_label("其他业务成本"): official_value("其他业务成本"),
        canonical_label("营业支出合计"): field("F036N"),
        canonical_label("利润总额"): field("F024N"),
        canonical_label("减：所得税费用"): field("F025N"),
        canonical_label("本行股东的净利润"): field("F028N"),
        canonical_label("少数股东的净利润"): field("F029N"),
        canonical_label("基本及稀释每股收益"): first_available(field("F031N"), field("F032N")),
        canonical_label("权益法下可转损益的其他综合收益"): official_value("权益法下可转损益的其他综合收益"),
        canonical_label("以公允价值计量且其变动计入其他综合收益的债务工具投资公允价值变动"): official_value(
            "以公允价值计量且其变动计入其他综合收益的债务工具投资公允价值变动"
        ),
        canonical_label("以公允价值计量且其变动计入其他综合收益的债务工具投资信用损失准备"): official_value(
            "以公允价值计量且其变动计入其他综合收益的债务工具投资信用损失准备"
        ),
        canonical_label("外币财务报表折算差额"): official_value("外币财务报表折算差额"),
        canonical_label("指定为以公允价值计量且其变动计入其他综合收益的权益工具投资公允价值变动"): official_value(
            "指定为以公允价值计量且其变动计入其他综合收益的权益工具投资公允价值变动"
        ),
        canonical_label("重新计量设定受益计划变动额"): official_value("重新计量设定受益计划变动额"),
        canonical_label("本年综合收益总额"): field("F039N"),
        canonical_label("本行股东的综合收益总额"): official_value("本行股东的综合收益总额"),
        canonical_label("少数股东的综合收益总额"): official_value("少数股东的综合收益总额"),
    }
)

BANK_CASH_RESOLVERS.update(
    {
        canonical_label("存放中央银行款项净减少额"): official_value("存放中央银行款项净减少额"),
        canonical_label("交易目的而持有的金融资产净减少额"): official_value("交易目的而持有的金融资产净减少额"),
        canonical_label("存放同业和其他金融机构款项净减少额"): official_value("存放同业和其他金融机构款项净减少额"),
        canonical_label("拆入资金及卖出回购金融资产款净增加额"): official_value("拆入资金及卖出回购金融资产款净增加额"),
        canonical_label("同业和其他金融机构存放款项净增加额"): official_value("同业和其他金融机构存放款项净增加额"),
        canonical_label("客户存款净增加额"): official_value("客户存款净增加额"),
        canonical_label("收到其他与经营活动有关的现金"): official_value("收到其他与经营活动有关的现金"),
        canonical_label("存放同业和其他金融机构款项净增加额"): official_value("存放同业和其他金融机构款项净增加额"),
        canonical_label("拆出资金及买入返售金融资产净增加额"): official_value("拆出资金及买入返售金融资产净增加额"),
        canonical_label("贷款和垫款净增加额"): field("F084N"),
        canonical_label("为交易目的而持有的金融资产净增加额"): official_value("为交易目的而持有的金融资产净增加额"),
        canonical_label("向中央银行借款净减少额"): official_value("向中央银行借款净减少额"),
        canonical_label("拆入资金及卖出回购金融资产款净减少额"): official_value("拆入资金及卖出回购金融资产款净减少额"),
        canonical_label("支付其他与经营活动有关的现金"): official_value("支付其他与经营活动有关的现金"),
        canonical_label("处置子公司、合营企业或联营企业收到的现金"): field("F019N"),
        canonical_label("出售固定资产和其他资产收到的现金"): field("F018N"),
        canonical_label("取得子公司、合营企业或联营企业支付的现金"): official_value("取得子公司、合营企业或联营企业支付的现金"),
        canonical_label("购建固定资产和其他资产所支付的现金"): field("F022N"),
        canonical_label("发行存款证及其他收到的现金"): official_value("发行存款证及其他收到的现金"),
        canonical_label("发行同业存单收到的现金"): official_value("发行同业存单收到的现金"),
        canonical_label("发行永续债募集的资金"): official_value("发行永续债募集的资金"),
        canonical_label("偿还存款证及其他支付的现金"): official_value("偿还存款证及其他支付的现金"),
        canonical_label("偿还同业存单支付的现金"): official_value("偿还同业存单支付的现金"),
        canonical_label("赎回永续债支付的现金"): official_value("赎回永续债支付的现金"),
        canonical_label("支付租赁负债的现金"): official_value("支付租赁负债的现金"),
        canonical_label("赎回永续债务资本利息支付的现金"): official_value("赎回永续债务资本利息支付的现金"),
        canonical_label("派发普通股股利支付的现金"): official_value("派发普通股股利支付的现金"),
        canonical_label("派发优先股股息支付的现金"): official_value("派发优先股股息支付的现金"),
        canonical_label("派发永续债利息支付的现金"): official_value("派发永续债利息支付的现金"),
        canonical_label("支付筹资活动的利息"): official_value("支付筹资活动的利息"),
        canonical_label("四、汇率变动对现金及现金等价物的影响额"): field("F037N"),
        canonical_label("因：汇率变动及现金及现金等价物的影响额"): field("F037N"),
        canonical_label("五、现金及现金等价物净（减少）/增加额"): field("F071N"),
        canonical_label("加：年初现金及现金等价物余额"): field("F040N"),
        canonical_label("六、年末现金及现金等价物余额"): field("F041N"),
    }
)

ROW_OCCURRENCE_RESOLVERS: dict[tuple[str, str], dict[tuple[str, int], Resolver]] = {
    ("company", "balance"): {
        (canonical_label("其他应收款"), 1): field("F011N"),
        (canonical_label("其他应收款"), 2): field("F014N"),
        (canonical_label("在建工程"), 1): sum_available_fields("F026N", "F027N"),
        (canonical_label("在建工程"), 2): aggregate_only("F026N", "F027N"),
        (canonical_label("其他应付款"), 1): field("F048N"),
        (canonical_label("其他应付款"), 2): subtract_fields("F048N", "F046N", "F047N"),
        (canonical_label("长期应付款"), 1): sum_available_fields("F056N", "F076N", "F077N"),
        (canonical_label("长期应付款"), 2): field("F076N"),
    },
}

TEMPLATE_LABEL_ALIASES: dict[tuple[str, str], dict[str, str]] = {
    ("company", "balance"): {
        canonical_label("发放贷款和垫款"): canonical_label("发放贷款及垫款"),
    },
    ("company", "income"): {
        canonical_label("一、营业总收入"): canonical_label("一、营业收入"),
        canonical_label("营业税金及附加"): canonical_label("税金及附加"),
    },
    ("company", "cash"): {
        canonical_label("收到的其他与筹资活动有关的现金"): canonical_label("收到其他与筹资活动有关的现金"),
        canonical_label("偿还债务所支付的现金"): canonical_label("偿还债务支付的现金"),
    },
    ("bank", "balance"): {
        canonical_label("衍生金融工具资产"): canonical_label("衍生金融资产"),
        canonical_label("买入返售金融资产"): canonical_label("买入返售款项"),
        canonical_label("衍生金融工具负债"): canonical_label("衍生金融负债"),
        canonical_label("卖出回购金融资产款"): canonical_label("卖出回购款项"),
    },
    ("bank", "income"): {
        canonical_label("投资净收益"): canonical_label("投资收益"),
        canonical_label("营业税金及附加"): canonical_label("税金及附加"),
        canonical_label("业务及管理费用"): canonical_label("业务及管理费"),
        canonical_label("其中：手续费及佣金收入"): canonical_label("其中:手续费及佣金收入"),
        canonical_label("其中：对联营公司的投资收益"): canonical_label("其中:对联营公司的投资收益"),
    },
    ("bank", "cash"): {
        canonical_label("收到的其他与筹资活动有关的现金"): canonical_label("收到其他与筹资活动有关的现金"),
        canonical_label("偿还债务所支付的现金"): canonical_label("偿还债务证券所支付的现金"),
    },
}

STATEMENT_RESOLVERS: dict[tuple[str, str], dict[str, Resolver]] = {
    ("company", "balance"): COMPANY_BALANCE_RESOLVERS,
    ("company", "income"): COMPANY_INCOME_RESOLVERS,
    ("company", "cash"): COMPANY_CASH_RESOLVERS,
    ("bank", "balance"): BANK_BALANCE_RESOLVERS,
    ("bank", "income"): BANK_INCOME_RESOLVERS,
    ("bank", "cash"): BANK_CASH_RESOLVERS,
}


MANUAL_FIELD_CATALOG: tuple[FieldCatalogItem, ...] = (
    FieldCatalogItem("F048N", "其他应付款合计", "company", "balance"),
    FieldCatalogItem("F056N", "长期应付款合计", "company", "balance"),
    FieldCatalogItem("F075N", "递延收益", "company", "balance"),
    FieldCatalogItem("F114N", "应付短期债券", "company", "balance"),
    FieldCatalogItem("F072N", "客户存款和同业存放款项净增加额", "bank", "cash"),
    FieldCatalogItem("F073N", "向中央银行借款净增加额", "bank", "cash"),
    FieldCatalogItem("F074N", "向其他金融机构拆入资金净增加额", "bank", "cash"),
    FieldCatalogItem("F081N", "收取利息、手续费及佣金的现金", "bank", "cash"),
    FieldCatalogItem("F084N", "客户贷款及垫款净增加额", "bank", "cash"),
    FieldCatalogItem("F085N", "存放中央银行和同业款项净增加额", "bank", "cash"),
    FieldCatalogItem("F087N", "支付利息、手续费及佣金的现金", "bank", "cash"),
    FieldCatalogItem("F103N", "其他权益工具", "bank", "balance"),
    FieldCatalogItem("F104N", "优先股", "bank", "balance"),
    FieldCatalogItem("F105N", "永续债", "bank", "balance"),
)


def build_api_field_catalog() -> dict[tuple[str, str], dict[str, str]]:
    catalog: dict[tuple[str, str], dict[str, str]] = {}
    for key, resolvers in STATEMENT_RESOLVERS.items():
        field_map = catalog.setdefault(key, {})
        for label_key, resolver in resolvers.items():
            if resolver_kind(resolver) != "field":
                continue
            source_fields = resolver_source_fields(resolver)
            if len(source_fields) != 1:
                continue
            field_map.setdefault(source_fields[0], label_key)

    for item in MANUAL_FIELD_CATALOG:
        catalog.setdefault((item.template_kind, item.statement_type), {})[item.field_name] = item.label
    return catalog


API_FIELD_CATALOG = build_api_field_catalog()


def build_supplemental_items() -> dict[tuple[str, str], tuple[SupplementalItem, ...]]:
    supplemental: dict[tuple[str, str], tuple[SupplementalItem, ...]] = {}
    for key, field_map in API_FIELD_CATALOG.items():
        ordered_items = tuple(
            SupplementalItem(label, field(field_name))
            for field_name, label in sorted(field_map.items(), key=lambda item: (item[1], item[0]))
        )
        supplemental[key] = ordered_items
    return supplemental


SUPPLEMENTAL_ITEMS = build_supplemental_items()


GENERIC_FIELD_LABELS = {
    field_name: label
    for label, field_name in DIRECT_FIELD_MAP.items()
}


def source_field_labels(template_kind: str, statement_type: str, source_fields: tuple[str, ...]) -> list[str]:
    field_map = API_FIELD_CATALOG.get((template_kind, statement_type), {})
    return [field_map.get(field_name, GENERIC_FIELD_LABELS.get(field_name, "接口汇总项")) for field_name in source_fields]


def describe_resolver(template_kind: str, statement_type: str, resolver: Resolver | None) -> str | None:
    if resolver is None:
        return None

    kind = resolver_kind(resolver)
    source_fields = resolver_source_fields(resolver)
    source_labels = resolver_source_labels(resolver)
    if kind == "official_sum":
        labels = [label for label in source_labels if label]
        return f"推导：{' + '.join(labels)}" if labels else "推导"
    if kind not in {"sum", "subtract", "sum_available"} or not source_fields:
        return None

    labels = source_field_labels(template_kind, statement_type, source_fields)
    if kind == "sum":
        return f"推导：{' + '.join(labels)}"
    if kind == "subtract":
        head, *tail = labels
        return f"推导：{head} - {' - '.join(tail)}"
    if kind == "sum_available":
        return f"推导：按有值项求和（{' + '.join(labels)}）"
    return None


@dataclass(frozen=True)
class SheetLayout:
    header_row: int
    unit_row: int | None
    data_start_row: int
    note_col: int
    data_start_col: int
    period_count: int


@dataclass(frozen=True)
class MissingRowExplanation:
    sheet_name: str
    row_index: int
    label: str
    category: str
    detail: str


def export_template_workbook(
    company: CompanyRecord,
    balance_records: list[dict],
    income_records: list[dict],
    cash_flow_records: list[dict],
    output_dir: str | Path,
    unit_label: str = DEFAULT_UNIT_LABEL,
    template_id: str | None = None,
    official_provider: "OfficialAnnualReportSource | None" = None,
) -> Path:
    template = resolve_template(template_id)
    normalized_unit_label = normalize_unit_label(unit_label)
    unit_scale = UNIT_SCALE_MAP[normalized_unit_label]
    workbook = load_workbook(template.path)
    output_path = ensure_writable_dir(output_dir)
    missing_rows: list[MissingRowExplanation] = []

    if MISSING_REASON_SHEET in workbook.sheetnames:
        del workbook[MISSING_REASON_SHEET]

    balance_periods = select_annual_records(balance_records, "ENDDATE")
    income_periods = select_annual_records(income_records, "F001D")
    cash_periods = select_annual_records(cash_flow_records, "F001D")

    for sheet in workbook.worksheets:
        statement_type = detect_statement_type(sheet.title)
        if statement_type is None:
            continue

        sheet.title = build_export_sheet_title(company.secname, STATEMENT_SHEET_SUFFIX[statement_type])
        base_font = detect_sheet_base_font(sheet)
        records = {
            "balance": balance_periods,
            "income": income_periods,
            "cash": cash_periods,
        }[statement_type]
        if not records:
            continue

        layout = prepare_sheet_layout(sheet, normalized_unit_label, len(records))
        periods = records[: layout.period_count]
        write_period_headers(sheet, layout, periods, normalized_unit_label)
        clear_period_cells(sheet, layout)
        normalize_template_labels(sheet, template, statement_type)
        attach_official_overrides(
            company=company,
            template=template,
            statement_type=statement_type,
            periods=periods,
            official_provider=official_provider,
            sheet=sheet,
            layout=layout,
        )
        sheet_missing_rows, covered_fields, template_label_keys = fill_statement_sheet(
            sheet=sheet,
            layout=layout,
            template=template,
            statement_type=statement_type,
            periods=periods,
            unit_scale=unit_scale,
        )
        missing_rows.extend(sheet_missing_rows)
        append_supplemental_section(
            sheet=sheet,
            layout=layout,
            template=template,
            statement_type=statement_type,
            periods=periods,
            unit_scale=unit_scale,
            covered_fields=covered_fields,
            template_label_keys=template_label_keys,
        )
        prune_statement_rows(sheet, template, statement_type)
        apply_consistent_fonts(sheet, base_font)
        apply_column_separators(sheet)
        auto_adjust_sheet_widths(sheet, layout)

    append_missing_reason_sheet(workbook, missing_rows, company.secname)

    latest_year = balance_periods[0][0][:4] if balance_periods else "latest"
    workbook_path = output_path / f"{company.secname}_{template.template_id}_{latest_year}YE.xlsx"
    workbook.save(workbook_path)
    return workbook_path


def attach_official_overrides(
    *,
    company: CompanyRecord,
    template: TemplateSpec,
    statement_type: str,
    periods: list[tuple[str, dict]],
    official_provider: "OfficialAnnualReportSource | None",
    sheet,
    layout: SheetLayout,
) -> None:
    if official_provider is None:
        return
    requested_labels = collect_requested_official_labels(sheet, layout)

    for period_end, record in periods:
        values = official_provider.get_statement_overrides(
            company,
            template_kind=template.kind,
            statement_type=statement_type,
            period_end=period_end,
            requested_labels=requested_labels,
        )
        if not values:
            continue
        official_rows = record.setdefault("__official_rows__", {})
        for label_entry, value in values.items():
            if isinstance(label_entry, tuple) and len(label_entry) == 2:
                label, occurrence = label_entry
            else:
                label = label_entry
                occurrence = 1
            for label_key in official_label_keys(label, template_kind=template.kind, statement_type=statement_type):
                official_rows[(label_key, occurrence)] = value
                official_rows.setdefault(label_key, value)


def collect_requested_official_labels(sheet, layout: SheetLayout) -> list[str]:
    labels: list[str] = []
    for row in range(layout.data_start_row, sheet.max_row + 1):
        value = sheet.cell(row, 1).value
        if value in (None, ""):
            continue
        labels.append(str(value).strip())
    return labels


def official_value_for_label(record: dict, label_key: str, occurrence: int = 1) -> object | None:
    official_rows = record.get("__official_rows__", {})
    if (label_key, occurrence) in official_rows:
        return official_rows[(label_key, occurrence)]
    if occurrence > 1:
        return None
    return official_rows.get(label_key)


def select_official_value(
    record: dict,
    label_key: str,
    occurrence: int,
    resolver: Resolver | None,
) -> object | None:
    official = official_value_for_label(record, label_key, occurrence)
    if official is None:
        return None
    if label_key not in OFFICIAL_OVERRIDE_GUARDED_LABELS or resolver is None or is_placeholder_resolver(resolver):
        return official

    api_value = resolver(record)
    if not isinstance(official, (int, float)) or not isinstance(api_value, (int, float)):
        return official
    if abs(official - api_value) > max(abs(api_value) * 0.1, 1.0):
        return None
    return official


def classify_row_source(
    template_kind: str,
    statement_type: str,
    label_key: str,
    occurrence: int,
    resolver: Resolver | None,
    periods: list[tuple[str, dict]],
) -> str | None:
    if resolver is None or is_placeholder_resolver(resolver):
        return None

    derived = is_derived_resolver(resolver)

    for _period, record in periods:
        official = select_official_value(record, label_key, occurrence, resolver)
        if official is not None:
            if has_pdf_api_conflict(label_key, occurrence, resolver, [(None, record)]):
                return "PDF年报（与API不一致）"
            return "PDF年报"

        value = resolver(record)
        if value is None:
            continue
        if derived:
            return describe_resolver(template_kind, statement_type, resolver) or "推导"
        return "API接口"
    return None


def has_pdf_api_conflict(
    label_key: str,
    occurrence: int,
    resolver: Resolver | None,
    periods: list[tuple[str | None, dict]],
) -> bool:
    if resolver is None or is_placeholder_resolver(resolver):
        return False

    for _period, record in periods:
        official = select_official_value(record, label_key, occurrence, resolver)
        api_value = resolver(record)
        if official is None or api_value is None:
            continue
        if values_conflict(official, api_value):
            return True
    return False


def values_conflict(left: object, right: object) -> bool:
    if isinstance(left, (int, float)) and isinstance(right, (int, float)):
        tolerance = max(1e-6, max(abs(left), abs(right)) * 1e-6)
        return abs(left - right) > tolerance
    return left != right


def select_annual_records(records: list[dict], date_key: str) -> list[tuple[str, dict]]:
    selected: list[tuple[str, dict]] = []
    for record in records:
        date_value = str(record.get(date_key, "")).strip()
        if not date_value.endswith("12-31"):
            continue
        if date_value != str(record.get("ENDDATE", "")).strip():
            continue
        if str(record.get("F002V", "")).strip() != "071001":
            continue
        selected.append((date_value, record))

    selected.sort(key=lambda item: item[0], reverse=True)
    return selected


def detect_statement_type(sheet_title: str) -> str | None:
    normalized = canonical_label(sheet_title)
    if "资产负债表" in normalized:
        return "balance"
    if "利润表" in normalized:
        return "income"
    if "现金流量表" in normalized:
        return "cash"
    return None


def build_export_sheet_title(company_name: str, suffix: str) -> str:
    return f"{company_name}{suffix}"[:31]


def prepare_sheet_layout(sheet, unit_label: str, record_count: int) -> SheetLayout:
    dual_header = uses_dual_header_rows(sheet)
    if dual_header:
        style_header_label_cell(sheet.cell(1, 1), "报表日期")
        style_header_label_cell(sheet.cell(2, 1), "单位")
        data_start_col, existing_count = detect_period_columns(sheet, 1, min_col=2)
        period_count = max(existing_count, record_count, DEFAULT_ANNUAL_PERIODS)
        note_col = ensure_note_column(sheet, target_col=data_start_col + period_count, dual_header=True)
        return SheetLayout(
            header_row=1,
            unit_row=2,
            data_start_row=3,
            note_col=note_col,
            data_start_col=data_start_col,
            period_count=period_count,
        )

    header_row = ensure_single_header_row(sheet, unit_label)
    data_start_col, existing_count = detect_period_columns(sheet, header_row, min_col=2)
    period_count = max(existing_count, record_count, DEFAULT_ANNUAL_PERIODS)
    note_col = ensure_note_column(sheet, target_col=data_start_col + period_count, dual_header=False)
    return SheetLayout(
        header_row=header_row,
        unit_row=None,
        data_start_row=header_row + 1,
        note_col=note_col,
        data_start_col=data_start_col,
        period_count=period_count,
    )


def uses_dual_header_rows(sheet) -> bool:
    return canonical_label(sheet.cell(1, 1).value) == canonical_label("报表日期") and canonical_label(
        sheet.cell(2, 1).value
    ) == canonical_label("单位")


def style_header_label_cell(cell, value: str) -> None:
    cell.value = value
    cell.alignment = Alignment(horizontal="left", vertical="center")
    cell.fill = HEADER_FILL
    cell.font = HEADER_FONT


def find_note_column(sheet) -> int | None:
    for column in range(2, sheet.max_column + 1):
        if canonical_label(sheet.cell(1, column).value) == canonical_label(NOTE_HEADER):
            return column
    return None


def ensure_note_column(sheet, *, target_col: int, dual_header: bool) -> int:
    note_col = max(target_col, 2)
    existing_col = find_note_column(sheet)
    if existing_col is None:
        sheet.insert_cols(note_col)
    elif existing_col < note_col:
        sheet.insert_cols(note_col + 1)
        sheet.delete_cols(existing_col)
    elif existing_col > note_col:
        sheet.insert_cols(note_col)
        sheet.delete_cols(existing_col + 1)

    style_header_label_cell(sheet.cell(1, note_col), NOTE_HEADER)
    if dual_header:
        style_header_label_cell(sheet.cell(2, note_col), NOTE_SUBHEADER)
    return note_col


def ensure_single_header_row(sheet, unit_label: str) -> int:
    first_cell_key = canonical_label(sheet.cell(1, 1).value)
    if first_cell_key not in {canonical_label("项目"), canonical_label("报表日期")}:
        sheet.insert_rows(1)
    title_cell = sheet.cell(1, 1)
    title_cell.value = f"项目（单位：{unit_label}）"
    title_cell.alignment = Alignment(horizontal="left", vertical="center")
    title_cell.fill = HEADER_FILL
    title_cell.font = HEADER_FONT
    return 1


def detect_period_columns(sheet, header_row: int, min_col: int = 2) -> tuple[int, int]:
    first_period_col: int | None = None
    existing_count = 0
    for column in range(min_col, sheet.max_column + 1):
        if looks_like_period_header(sheet.cell(header_row, column).value):
            first_period_col = column
            break

    if first_period_col is not None:
        column = first_period_col
        while column <= sheet.max_column and looks_like_period_header(sheet.cell(header_row, column).value):
            existing_count += 1
            column += 1
        return first_period_col, existing_count

    second_cell = str(sheet.cell(header_row, min_col).value or "")
    if "附注" in second_cell:
        return min_col + 1, 0
    return min_col, 0


def looks_like_period_header(value: object | None) -> bool:
    if isinstance(value, (date,)):
        return True
    if isinstance(value, (int, float)):
        return value > 10_000
    if isinstance(value, str):
        text = value.strip()
        return bool(DATE_PATTERN.search(text) or re.fullmatch(r"\d{4}(?:-\d{2}-\d{2})?", text))
    return False


def detect_sheet_base_font(sheet) -> Font:
    for row in range(1, sheet.max_row + 1):
        for column in range(1, sheet.max_column + 1):
            cell = sheet.cell(row, column)
            if cell.value not in (None, "") and cell.font is not None:
                return copy(cell.font)
    return Font(name="等线", sz=12)


def apply_consistent_fonts(sheet, base_font: Font) -> None:
    for row in range(1, sheet.max_row + 1):
        for column in range(1, sheet.max_column + 1):
            cell = sheet.cell(row, column)
            font = copy(cell.font)
            font.name = base_font.name
            font.sz = base_font.sz
            cell.font = font


def apply_column_separators(sheet) -> None:
    max_column = sheet.max_column
    for row in range(1, sheet.max_row + 1):
        for column in range(1, max_column + 1):
            cell = sheet.cell(row, column)
            border = copy(cell.border)
            border.left = SEPARATOR_SIDE
            if column == max_column:
                border.right = SEPARATOR_SIDE
            cell.border = border


def write_period_headers(
    sheet,
    layout: SheetLayout,
    periods: list[tuple[str, dict]],
    unit_label: str,
) -> None:
    for offset in range(layout.period_count):
        column = layout.data_start_col + offset
        cell = sheet.cell(layout.header_row, column)
        cell.alignment = Alignment(horizontal="center", vertical="center")
        cell.fill = HEADER_FILL
        cell.font = HEADER_FONT
        if offset < len(periods):
            period_date = date.fromisoformat(periods[offset][0])
            cell.value = period_date
            cell.number_format = "yyyy-mm-dd"
        else:
            cell.value = None

        if layout.unit_row is not None:
            unit_cell = sheet.cell(layout.unit_row, column)
            unit_cell.alignment = Alignment(horizontal="center", vertical="center")
            unit_cell.fill = HEADER_FILL
            unit_cell.font = HEADER_FONT
            unit_cell.value = unit_label


def clear_period_cells(sheet, layout: SheetLayout) -> None:
    final_col = layout.data_start_col + layout.period_count - 1
    for row in range(layout.data_start_row, sheet.max_row + 1):
        sheet.cell(row, layout.note_col).value = None
        for column in range(layout.data_start_col, final_col + 1):
            sheet.cell(row, column).value = None


def normalize_template_labels(sheet, template: TemplateSpec, statement_type: str) -> None:
    return


def prune_statement_rows(sheet, template: TemplateSpec, statement_type: str) -> None:
    if template.kind == "company" and statement_type == "income":
        for row in range(sheet.max_row, 0, -1):
            if canonical_label(sheet.cell(row, 1).value) == canonical_label("营业收入"):
                sheet.delete_rows(row, 1)

    for row in range(sheet.max_row, 0, -1):
        if all(sheet.cell(row, column).value in (None, "") for column in range(1, sheet.max_column + 1)):
            sheet.delete_rows(row, 1)


def is_placeholder_resolver(resolver: Resolver | None) -> bool:
    return resolver_kind(resolver) in {"aggregate_only", "formula_required", "no_api_field"}


def resolve_row_resolver(template_kind: str, statement_type: str, label_key: str, occurrence: int) -> Resolver | None:
    occurrence_map = ROW_OCCURRENCE_RESOLVERS.get((template_kind, statement_type), {})
    resolver = occurrence_map.get((label_key, occurrence))
    if resolver is not None:
        return resolver

    resolvers = STATEMENT_RESOLVERS[(template_kind, statement_type)]
    resolver = resolvers.get(label_key)
    if resolver is not None:
        return resolver

    alias_map = TEMPLATE_LABEL_ALIASES.get((template_kind, statement_type), {})
    aliased_key = alias_map.get(label_key)
    if aliased_key is None:
        return None
    return resolvers.get(aliased_key)


def write_resolved_value(
    cell,
    value: object | None,
    unit_scale: int,
    label_key: str,
    *,
    derived: bool = False,
) -> bool:
    if value is None:
        return False

    if isinstance(value, (int, float)) and label_key not in NON_SCALED_LABELS:
        cell.value = value / unit_scale
        cell.number_format = "#,##0"
    else:
        cell.value = value
        if label_key in NON_SCALED_LABELS:
            cell.number_format = "0.0000"
    cell.alignment = Alignment(horizontal="right", vertical="center")
    if derived:
        cell.fill = DERIVED_FILL
    return True


def find_last_label_row(sheet) -> int:
    for row in range(sheet.max_row, 0, -1):
        if sheet.cell(row, 1).value not in (None, ""):
            return row
    return sheet.max_row


def append_supplemental_section(
    sheet,
    layout: SheetLayout,
    template: TemplateSpec,
    statement_type: str,
    periods: list[tuple[str, dict]],
    unit_scale: int,
    covered_fields: set[str],
    template_label_keys: set[str],
) -> None:
    supplemental_items = SUPPLEMENTAL_ITEMS.get((template.kind, statement_type), ())
    rows_to_append: list[tuple[str, str, Resolver, list[object | None]]] = []

    for item in supplemental_items:
        label_key = canonical_label(item.label)
        if label_key in template_label_keys:
            continue

        source_fields = resolver_source_fields(item.resolver)
        if source_fields and all(field_name in covered_fields for field_name in source_fields):
            continue

        values = [item.resolver(record) for _, record in periods]
        if not any(value is not None for value in values):
            continue
        rows_to_append.append((item.label, label_key, item.resolver, values))

    if not rows_to_append:
        return

    title_row = find_last_label_row(sheet) + 2
    sheet.cell(title_row, 1, "补充项目")
    apply_section_style(sheet, title_row, statement_type, "补充项目", None)

    for offset, (label, label_key, resolver, values) in enumerate(rows_to_append, start=1):
        row_index = title_row + offset
        label_cell = sheet.cell(row_index, 1, label)
        label_cell.alignment = Alignment(horizontal="left", vertical="center", wrap_text=True)
        note_text = classify_row_source(template.kind, statement_type, label_key, 1, resolver, periods)
        note_cell = sheet.cell(row_index, layout.note_col)
        note_cell.value = note_text
        note_cell.alignment = Alignment(horizontal="left", vertical="center", wrap_text=True)
        if isinstance(note_text, str) and note_text.startswith("推导"):
            note_cell.fill = DERIVED_FILL
        for period_index, value in enumerate(values):
            cell = sheet.cell(row_index, layout.data_start_col + period_index)
            write_resolved_value(
                cell,
                value,
                unit_scale,
                label_key,
                derived=is_derived_resolver(resolver),
            )


def fill_statement_sheet(
    sheet,
    layout: SheetLayout,
    template: TemplateSpec,
    statement_type: str,
    periods: list[tuple[str, dict]],
    unit_scale: int,
) -> tuple[list[MissingRowExplanation], set[str], set[str]]:
    explanations: list[MissingRowExplanation] = []
    covered_fields: set[str] = set()
    template_label_keys: set[str] = set()
    occurrences: dict[str, int] = {}

    for row in range(layout.data_start_row, sheet.max_row + 1):
        label_cell = sheet.cell(row, 1)
        raw_label = label_cell.value
        if raw_label in (None, ""):
            continue
        label_cell.alignment = Alignment(horizontal="left", vertical="center", wrap_text=True)
        label_key = canonical_label(raw_label)
        template_label_keys.add(label_key)
        occurrences[label_key] = occurrences.get(label_key, 0) + 1
        resolver = resolve_row_resolver(template.kind, statement_type, label_key, occurrences[label_key])
        apply_section_style(sheet, row, statement_type, raw_label, resolver)
        note_cell = sheet.cell(row, layout.note_col)
        note_text = classify_row_source(
            template.kind,
            statement_type,
            label_key,
            occurrences[label_key],
            resolver,
            periods,
        )
        note_cell.value = note_text
        note_cell.alignment = Alignment(horizontal="left", vertical="center", wrap_text=True)
        if isinstance(note_text, str) and note_text.startswith("推导"):
            note_cell.fill = DERIVED_FILL
        if resolver is not None and not is_placeholder_resolver(resolver):
            covered_fields.update(resolver_source_fields(resolver))
        if resolver is None:
            explanation = build_missing_row_explanation(
                sheet_name=sheet.title,
                row_index=row,
                raw_label=raw_label,
                resolver=None,
                template_kind=template.kind,
                statement_type=statement_type,
                periods=periods,
            )
            if explanation is not None:
                explanations.append(explanation)
            continue

        resolved_any_value = False
        for offset, (_period, record) in enumerate(periods):
            official_value = select_official_value(record, label_key, occurrences[label_key], resolver)
            used_official = official_value is not None
            value = official_value if used_official else resolver(record)
            cell = sheet.cell(row, layout.data_start_col + offset)
            resolved_any_value = (
                write_resolved_value(
                    cell,
                    value,
                    unit_scale,
                    label_key,
                    derived=is_derived_resolver(resolver) and not used_official,
                )
                or resolved_any_value
            )
        if not resolved_any_value:
            explanation = build_missing_row_explanation(
                sheet_name=sheet.title,
                row_index=row,
                raw_label=raw_label,
                resolver=resolver,
                template_kind=template.kind,
                statement_type=statement_type,
                periods=periods,
            )
            if explanation is not None:
                explanations.append(explanation)

    return explanations, covered_fields, template_label_keys


def apply_section_style(
    sheet,
    row_index: int,
    statement_type: str,
    raw_label: object | None,
    resolver: Resolver | None,
) -> None:
    if not is_section_style_label(raw_label, resolver):
        return

    fill = SECTION_FILL_BY_STATEMENT[statement_type]
    for column in range(1, sheet.max_column + 1):
        cell = sheet.cell(row_index, column)
        cell.fill = fill
        cell.font = SECTION_FONT
        if column == 1:
            cell.alignment = Alignment(horizontal="left", vertical="center", wrap_text=True)


def is_numbered_or_exact_section_label(raw_label: object | None) -> bool:
    text = str(raw_label or "").strip()
    if not text:
        return False
    if re.match(r"^[（(]\d+[)）]", text):
        return False
    if re.match(r"^\d+[、.．)]", text):
        return False
    normalized_text = text.rstrip("：:；;")
    if normalized_text in SECTION_LABEL_EXACT:
        return True
    if re.match(r"^[一二三四五六七八九十]+[、.．]", text):
        return True
    return False


def is_section_style_label(raw_label: object | None, resolver: Resolver | None) -> bool:
    if is_numbered_or_exact_section_label(raw_label):
        return True

    text = str(raw_label or "").strip()
    normalized_text = text.rstrip("：:；;")
    if resolver is None and text.endswith(("：", ":", "；", ";")) and len(normalized_text) <= 20:
        return True
    return False


def is_section_label(raw_label: object | None, resolver: Resolver | None) -> bool:
    if resolver is not None:
        return False
    return is_section_style_label(raw_label, resolver)


def build_missing_row_explanation(
    *,
    sheet_name: str,
    row_index: int,
    raw_label: object | None,
    resolver: Resolver | None,
    template_kind: str,
    statement_type: str,
    periods: list[tuple[str, dict]],
) -> MissingRowExplanation | None:
    label = str(raw_label or "").strip()
    if not label:
        return None

    if resolver is None:
        if is_section_label(raw_label, resolver):
            return MissingRowExplanation(
                sheet_name=sheet_name,
                row_index=row_index,
                label=label,
                category="模板分组行",
                detail="该行用于版式分组，不对应具体取数项目。",
            )
        return MissingRowExplanation(
            sheet_name=sheet_name,
            row_index=row_index,
            label=label,
            category="模板有此行但接口无对应字段",
            detail="模板里有该项目，但当前接口字典中没有可稳定对应的字段。",
        )

    resolver_type = resolver_kind(resolver)
    source_fields = resolver_source_fields(resolver)
    source_labels = resolver_source_labels(resolver)
    if resolver_type == "official":
        labels = "、".join(source_labels) if source_labels else label
        return MissingRowExplanation(
            sheet_name=sheet_name,
            row_index=row_index,
            label=label,
            category="官网年报未取到稳定值",
            detail=f"这行已切换为官网年报取值：{labels}。当前导出的所有年度里，没有从官网年报表格中稳定提取到可写入的数值。",
        )

    if resolver_type == "official_sum":
        labels = "、".join(source_labels) if source_labels else label
        return MissingRowExplanation(
            sheet_name=sheet_name,
            row_index=row_index,
            label=label,
            category="官网年报可用项不足，无法安全推导",
            detail=f"这行依赖官网年报项目推导：{labels}。当前导出的所有年度里，组成项不足，不能安全计算。",
        )
    if resolver_type == "no_api_field":
        return MissingRowExplanation(
            sheet_name=sheet_name,
            row_index=row_index,
            label=label,
            category="模板有此行但接口无对应字段",
            detail="模板里有该项目，但当前接口没有可直接对应的独立字段。",
        )

    if not source_fields:
        return MissingRowExplanation(
            sheet_name=sheet_name,
            row_index=row_index,
            label=label,
            category="需要公式但当前未实现",
            detail="这行需要额外公式或口径转换，目前还没有安全可用的计算规则。",
        )

    present_fields = [field_name for field_name in source_fields if any(field_name in record for _, record in periods)]
    present_field_labels = source_field_labels(template_kind, statement_type, tuple(present_fields))
    if not present_fields:
        return MissingRowExplanation(
            sheet_name=sheet_name,
            row_index=row_index,
            label=label,
            category="模板有此行但接口无对应字段",
            detail="当前已接入接口返回里没有这类独立字段。",
        )

    non_empty_fields = [
        field_name
        for field_name in source_fields
        if any(record.get(field_name) not in (None, "") for _, record in periods)
    ]
    non_empty_field_labels = source_field_labels(template_kind, statement_type, tuple(non_empty_fields))
    if not non_empty_fields:
        return MissingRowExplanation(
            sheet_name=sheet_name,
            row_index=row_index,
            label=label,
            category="接口字段存在但所选年度无值",
            detail=f"当前已接入接口里能看到对应字段，但这次导出的所有年度都没有披露值：{', '.join(present_field_labels)}。",
        )

    if resolver_type == "aggregate_only":
        return MissingRowExplanation(
            sheet_name=sheet_name,
            row_index=row_index,
            label=label,
            category="接口只有汇总值、没有拆分字段",
            detail=f"在当前已接入并核过的接口范围里，只找到了汇总字段：{', '.join(non_empty_field_labels)}；没有足够稳定的拆分字段可安全填到这一行。",
        )

    if resolver_type == "formula_required":
        return MissingRowExplanation(
            sheet_name=sheet_name,
            row_index=row_index,
            label=label,
            category="需要公式但当前未实现",
            detail=f"接口里已有相关字段：{', '.join(non_empty_field_labels)}，但还缺安全的计算公式。",
        )

    return MissingRowExplanation(
        sheet_name=sheet_name,
        row_index=row_index,
        label=label,
        category="需要公式但当前未实现",
        detail=f"已找到来源字段：{', '.join(non_empty_field_labels)}，但当前规则还不能安全产出结果。",
    )


def append_missing_reason_sheet(workbook, explanations: list[MissingRowExplanation], company_name: str) -> None:
    if not explanations:
        return

    sheet = workbook.create_sheet(build_export_sheet_title(company_name, MISSING_REASON_SHEET))
    headers = ("工作表", "行号", "项目", "原因分类", "说明")
    for column, header in enumerate(headers, start=1):
        cell = sheet.cell(1, column, header)
        cell.fill = HEADER_FILL
        cell.font = HEADER_FONT
        cell.alignment = Alignment(horizontal="center", vertical="center")

    for row_index, explanation in enumerate(explanations, start=2):
        sheet.cell(row_index, 1, explanation.sheet_name)
        sheet.cell(row_index, 2, explanation.row_index)
        sheet.cell(row_index, 3, explanation.label)
        sheet.cell(row_index, 4, explanation.category)
        sheet.cell(row_index, 5, explanation.detail)

    auto_adjust_sheet_widths(sheet)


def auto_adjust_sheet_widths(sheet, layout: SheetLayout | None = None) -> None:
    max_column = sheet.max_column
    for column in range(1, max_column + 1):
        letter = get_column_letter(column)
        max_width = 0
        for row in range(1, sheet.max_row + 1):
            cell = sheet.cell(row, column)
            max_width = max(max_width, display_text_width(cell.value, cell.number_format))

        padding = 3 if column == 1 else 2
        width = max_width + padding
        if column == 1:
            width = min(max(width, 18), 28)
        if layout is not None and column == layout.note_col:
            width = min(max(width, 18), 30)
        if layout is not None and column >= layout.data_start_col:
            width = max(width, 12)
        if column == 1:
            sheet.column_dimensions[letter].width = width
        else:
            sheet.column_dimensions[letter].width = min(width, 42)

    wrap_label_and_note_columns_and_adjust_row_heights(sheet, layout)


def wrap_label_and_note_columns_and_adjust_row_heights(sheet, layout: SheetLayout | None = None) -> None:
    first_col_width = float(sheet.column_dimensions["A"].width or 18)
    usable_width = max(first_col_width - 2, 8)
    note_col_width = None
    note_letter = None
    if layout is not None:
        note_letter = get_column_letter(layout.note_col)
        note_col_width = max(float(sheet.column_dimensions[note_letter].width or 18) - 2, 10)

    for row in range(1, sheet.max_row + 1):
        cell = sheet.cell(row, 1)
        cell.alignment = Alignment(horizontal="left", vertical="center", wrap_text=True)
        text_width = display_text_width(cell.value, cell.number_format)
        line_count = 1
        if text_width > 0:
            line_count = max(line_count, int((text_width + usable_width - 1) // usable_width))

        if layout is not None:
            note_cell = sheet.cell(row, layout.note_col)
            note_cell.alignment = Alignment(horizontal="left", vertical="center", wrap_text=True)
            note_width = display_text_width(note_cell.value, note_cell.number_format)
            if note_width > 0 and note_col_width is not None:
                line_count = max(line_count, int((note_width + note_col_width - 1) // note_col_width))

        minimum_height = 22 if row == 1 else 20
        sheet.row_dimensions[row].height = max(minimum_height, 18 * line_count)


def display_text_width(value: object | None, number_format: str | None = None) -> int:
    if value is None:
        return 0
    if isinstance(value, date):
        text = value.isoformat()
    elif isinstance(value, (int, float)) and number_format == "#,##0":
        text = f"{value:,.0f}"
    else:
        text = str(value)
    return sum(2 if unicodedata.east_asian_width(char) in {"W", "F"} else 1 for char in text)
