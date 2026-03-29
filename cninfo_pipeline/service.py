from __future__ import annotations

import shutil
from dataclasses import dataclass
from pathlib import Path
from typing import Callable

from openpyxl import Workbook
from openpyxl.styles import Alignment, Font

from .client import CninfoClient, CompanyRecord


ProgressCallback = Callable[[int, str], None]
CACHE_DIR_NAME = ".cninfo_internal"

ROW_ORDER = [
    "报表日期",
    "单位",
    "流动资产",
    "货币资金",
    "交易性金融资产",
    "衍生金融资产",
    "应收票据及应收账款8+9",
    "应收票据",
    "应收账款",
    "应收款项融资",
    "预付款项",
    "其他应收款(合计)13+14+15",
    "应收利息",
    "应收股利",
    "其他应收款",
    "买入返售金融资产",
    "存货",
    "划分为持有待售的资产（合同资产）",
    "一年内到期的非流动资产",
    "待摊费用",
    "待处理流动资产损益",
    "其他流动资产",
    "流动资产合计",
    "非流动资产",
    "发放贷款及垫款",
    "可供出售金融资产（其他权益工具投资）",
    "持有至到期投资（其他非流动金融资产）",
    "长期应收款（债权投资）",
    "长期股权投资",
    "投资性房地产",
    "在建工程(合计)32+33",
    "在建工程",
    "工程物资",
    "固定资产及清理(合计)35+36",
    "固定资产净额",
    "固定资产清理",
    "生产性生物资产",
    "公益性生物资产",
    "油气资产",
    "使用权资产",
    "无形资产",
    "开发支出",
    "商誉",
    "长期待摊费用",
    "递延所得税资产",
    "其他非流动资产",
    "非流动资产合计",
    "资产总计",
    "流动负债",
    "短期借款",
    "交易性金融负债",
    "应付票据及应付账款53+54",
    "应付票据",
    "应付账款",
    "预收款项",
    "应付手续费及佣金（合同负债）",
    "应付职工薪酬",
    "应交税费",
    "其他应付款(合计)60+61+62",
    "应付利息",
    "应付股利",
    "其他应付款",
    "预提费用",
    "一年内的递延收益",
    "应付短期债券",
    "一年内到期的非流动负债",
    "其他流动负债",
    "流动负债合计",
    "非流动负债",
    "长期借款",
    "应付债券",
    "租赁负债",
    "长期应付职工薪酬",
    "长期应付款(合计)75+76",
    "长期应付款",
    "专项应付款",
    "预计非流动负债",
    "递延所得税负债",
    "长期递延收益",
    "其他非流动负债",
    "非流动负债合计",
    "负债合计",
    "所有者权益",
    "实收资本(或股本)",
    "资本公积",
    "减：库存股",
    "其他综合收益",
    "专项储备",
    "盈余公积",
    "一般风险准备",
    "未分配利润",
    "归属于母公司股东权益合计",
    "少数股东权益",
    "所有者权益(或股东权益)合计",
    "负债和所有者权益(或股东权益)总计",
    "",
    "实际资产负债率",
    "FA&CIP占比资产总额",
    "商誉占比归属股东权益",
    "FA&CIP净额",
    "FA&CIP净额-合计1",
    "FA&CIP净额-合计2",
    "权益乘数=资产总额/所有者权益",
    "产权比率=负债总额/归属股东权益",
]

SECTION_ROWS = {
    "流动资产",
    "非流动资产",
    "流动负债",
    "非流动负债",
    "所有者权益",
}

RATIO_ROWS = {
    "实际资产负债率",
    "FA&CIP占比资产总额",
    "商誉占比归属股东权益",
    "权益乘数=资产总额/所有者权益",
    "产权比率=负债总额/归属股东权益",
}

DIRECT_FIELD_MAP = {
    "货币资金": "F006N",
    "交易性金融资产": "F117N",
    "衍生金融资产": "F080N",
    "应收票据": "F008N",
    "应收账款": "F009N",
    "应收款项融资": "F110N",
    "预付款项": "F010N",
    "其他应收款(合计)13+14+15": "F011N",
    "应收利息": "F012N",
    "应收股利": "F013N",
    "其他应收款": "F014N",
    "买入返售金融资产": "F016N",
    "存货": "F015N",
    "划分为持有待售的资产（合同资产）": "F119N",
    "一年内到期的非流动资产": "F017N",
    "待摊费用": "F020N",
    "待处理流动资产损益": "F021N",
    "其他流动资产": "F018N",
    "流动资产合计": "F019N",
    "发放贷款及垫款": "F113N",
    "可供出售金融资产（其他权益工具投资）": "F111N",
    "持有至到期投资（其他非流动金融资产）": "F112N",
    "长期应收款（债权投资）": "F022N",
    "长期股权投资": "F023N",
    "投资性房地产": "F024N",
    "在建工程": "F026N",
    "工程物资": "F027N",
    "固定资产净额": "F025N",
    "固定资产清理": "F028N",
    "生产性生物资产": "F029N",
    "公益性生物资产": "F030N",
    "油气资产": "F040N",
    "使用权资产": "F121N",
    "无形资产": "F031N",
    "开发支出": "F032N",
    "商誉": "F033N",
    "长期待摊费用": "F034N",
    "递延所得税资产": "F035N",
    "其他非流动资产": "F036N",
    "非流动资产合计": "F037N",
    "资产总计": "F038N",
    "短期借款": "F039N",
    "交易性金融负债": "F090N",
    "应付票据": "F041N",
    "应付账款": "F042N",
    "预收款项": "F043N",
    "应付手续费及佣金（合同负债）": "F115N",
    "应付职工薪酬": "F044N",
    "应交税费": "F045N",
    "应付利息": "F046N",
    "应付股利": "F047N",
    "预提费用": "F049N",
    "一年内的递延收益": "F116N",
    "应付短期债券": "F114N",
    "一年内到期的非流动负债": "F050N",
    "其他流动负债": "F051N",
    "流动负债合计": "F052N",
    "长期借款": "F053N",
    "应付债券": "F054N",
    "租赁负债": "F122N",
    "长期应付职工薪酬": "F055N",
    "长期应付款(合计)75+76": "F056N",
    "长期应付款": "F076N",
    "专项应付款": "F077N",
    "预计非流动负债": "F057N",
    "递延所得税负债": "F058N",
    "长期递延收益": "F075N",
    "其他非流动负债": "F059N",
    "非流动负债合计": "F060N",
    "负债合计": "F061N",
    "实收资本(或股本)": "F062N",
    "资本公积": "F063N",
    "减：库存股": "F066N",
    "其他综合收益": "F074N",
    "专项储备": "F068N",
    "盈余公积": "F064N",
    "一般风险准备": "F069N",
    "未分配利润": "F065N",
    "归属于母公司股东权益合计": "F073N",
    "少数股东权益": "F067N",
    "所有者权益(或股东权益)合计": "F070N",
    "负债和所有者权益(或股东权益)总计": "F071N",
}


@dataclass(frozen=True)
class PipelineResult:
    company: CompanyRecord
    output_path: Path
    total_records: int
    annual_records: int


class AnnualReportPipeline:
    def __init__(self, client: CninfoClient | None = None) -> None:
        self.client = client or CninfoClient()

    def run(
        self,
        company_query: str = "长江电力",
        output_dir: str | Path = "outputs",
        progress: ProgressCallback | None = None,
    ) -> PipelineResult:
        reporter = progress or (lambda _percent, _message: None)
        cache_dir = prepare_cache_dir(Path(output_dir))
        self.client.set_cache_dir(cache_dir)

        reporter(10, "正在匹配公司信息...")
        company = self.client.search_company(company_query)

        reporter(35, f"正在抓取 {company.secname} 的资产负债表...")
        records = self.client.fetch_balance_sheet(company.seccode)

        reporter(60, "正在筛选年报（仅保留 12-31 合并本期）...")
        annual_records = filter_annual_merged_records(records)

        reporter(80, "正在整理为模板报表格式...")
        matrix = build_statement_matrix(annual_records)

        reporter(90, "正在导出 Excel...")
        output_path = export_statement_workbook(company, matrix, output_dir)

        reporter(100, f"完成，已导出 {len(annual_records)} 份年报到 {output_path.name}")
        return PipelineResult(
            company=company,
            output_path=output_path,
            total_records=len(records),
            annual_records=len(annual_records),
        )


def filter_annual_merged_records(records: list[dict]) -> list[dict]:
    annual = [
        record
        for record in records
        if str(record.get("ENDDATE", "")).endswith("12-31")
        and record.get("F003V") == "合并本期"
    ]
    annual.sort(key=lambda item: str(item.get("ENDDATE", "")), reverse=True)
    return annual


def build_statement_matrix(records: list[dict]) -> list[list[object | None]]:
    date_keys = [str(record["ENDDATE"]).replace("-", "") for record in records]
    rows: list[list[object | None]] = []

    for label in ROW_ORDER:
        if label == "报表日期":
            rows.append([label, *[int(date_key) for date_key in date_keys]])
            continue
        if label == "单位":
            rows.append([label, *(["元"] * len(date_keys))])
            continue
        if label in SECTION_ROWS or label == "":
            rows.append([label, *([None] * len(date_keys))])
            continue

        row_values = [compute_row_value(label, record) for record in records]
        rows.append([label, *row_values])

    return rows


def compute_row_value(label: str, record: dict) -> object | None:
    if label in DIRECT_FIELD_MAP:
        return record.get(DIRECT_FIELD_MAP[label])

    if label == "应收票据及应收账款8+9":
        return add_values(record, "应收票据", "应收账款")
    if label == "在建工程(合计)32+33":
        direct = record.get("F026N")
        return direct if direct is not None else add_values(record, "在建工程", "工程物资")
    if label == "固定资产及清理(合计)35+36":
        direct = record.get("F025N")
        return direct if direct is not None else add_values(record, "固定资产净额", "固定资产清理")
    if label == "应付票据及应付账款53+54":
        return add_values(record, "应付票据", "应付账款")
    if label == "其他应付款(合计)60+61+62":
        direct = record.get("F048N")
        return direct if direct is not None else add_values(record, "应付利息", "应付股利", "其他应付款")
    if label == "其他应付款":
        total = record.get("F048N")
        if total is None:
            return None
        interest = numeric_or_zero(record.get("F046N"))
        dividend = numeric_or_zero(record.get("F047N"))
        return total - interest - dividend
    if label == "FA&CIP净额":
        return add_values(record, "在建工程(合计)32+33", "固定资产及清理(合计)35+36")
    if label == "FA&CIP净额-合计1":
        return compute_row_value("固定资产及清理(合计)35+36", record)
    if label == "FA&CIP净额-合计2":
        return compute_row_value("FA&CIP净额", record)
    if label == "实际资产负债率":
        return safe_divide(
            compute_row_value("负债合计", record),
            compute_row_value("资产总计", record),
        )
    if label == "FA&CIP占比资产总额":
        return safe_divide(
            compute_row_value("FA&CIP净额", record),
            compute_row_value("资产总计", record),
        )
    if label == "商誉占比归属股东权益":
        return safe_divide(
            compute_row_value("商誉", record),
            compute_row_value("归属于母公司股东权益合计", record),
        )
    if label == "权益乘数=资产总额/所有者权益":
        return safe_divide(
            compute_row_value("资产总计", record),
            compute_row_value("所有者权益(或股东权益)合计", record),
        )
    if label == "产权比率=负债总额/归属股东权益":
        return safe_divide(
            compute_row_value("负债合计", record),
            compute_row_value("归属于母公司股东权益合计", record),
        )

    return None


def add_values(record: dict, *labels: str) -> float | None:
    values = [compute_row_value(label, record) for label in labels]
    numeric = [value for value in values if isinstance(value, (int, float))]
    return sum(numeric) if numeric else None


def numeric_or_zero(value: object | None) -> float:
    return float(value) if isinstance(value, (int, float)) else 0.0


def safe_divide(numerator: object | None, denominator: object | None) -> float | None:
    if not isinstance(numerator, (int, float)) or not isinstance(denominator, (int, float)):
        return None
    if denominator == 0:
        return None
    return numerator / denominator


def prepare_cache_dir(output_dir: Path) -> Path:
    cache_dir = output_dir / CACHE_DIR_NAME
    legacy_cache_dir = output_dir / ".cache"

    if legacy_cache_dir.exists() and not cache_dir.exists():
        shutil.move(str(legacy_cache_dir), str(cache_dir))
    elif legacy_cache_dir.exists() and cache_dir.exists():
        shutil.rmtree(legacy_cache_dir, ignore_errors=True)

    return cache_dir


def export_statement_workbook(
    company: CompanyRecord,
    matrix: list[list[object | None]],
    output_dir: str | Path,
) -> Path:
    output_path = Path(output_dir)
    output_path.mkdir(parents=True, exist_ok=True)

    workbook = Workbook()
    sheet = workbook.active
    sheet.title = f"{company.secname}资产负债表"
    regular_font = Font(name="等线", size=11)
    bold_font = Font(name="等线", size=11, bold=True)

    for row_index, row in enumerate(matrix, start=1):
        for column_index, value in enumerate(row, start=1):
            cell = sheet.cell(row=row_index, column=column_index, value=value)
            cell.font = regular_font
            if column_index == 1:
                cell.alignment = Alignment(horizontal="left", vertical="center")
            else:
                cell.alignment = Alignment(horizontal="center", vertical="center")
                if row[0] == "报表日期":
                    cell.number_format = "0"
                elif row[0] in RATIO_ROWS and isinstance(value, (int, float)):
                    cell.number_format = "0.00%"

        label = row[0]
        if label in SECTION_ROWS:
            for column_index in range(1, len(row) + 1):
                sheet.cell(row=row_index, column=column_index).font = bold_font
        if label in {"报表日期", "单位"}:
            for column_index in range(1, len(row) + 1):
                sheet.cell(row=row_index, column=column_index).font = bold_font

    sheet.column_dimensions["A"].width = 34
    for column_index in range(2, len(matrix[0]) + 1):
        column_letter = sheet.cell(row=1, column=column_index).column_letter
        sheet.column_dimensions[column_letter].width = 16
    sheet.freeze_panes = "B3"

    safe_name = f"{company.seccode}_{company.secname}_annual_balance_sheet.xlsx"
    workbook_path = output_path / safe_name
    workbook.save(workbook_path)
    return workbook_path
