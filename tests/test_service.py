from __future__ import annotations

import shutil
import tempfile
import unittest
from pathlib import Path
from unittest.mock import MagicMock, patch

from openpyxl import Workbook, load_workbook

from cninfo_pipeline.client import CompanyRecord
from cninfo_pipeline.official_source import (
    extract_bank_balance_values_from_text,
    extract_company_income_values_from_text,
    extract_statement_values_from_text,
)
from cninfo_pipeline.service import (
    AnnualReportPipeline,
    build_statement_matrix,
    export_statement_workbook,
    filter_annual_merged_records,
    prepare_cache_dir,
)
from cninfo_pipeline.template_export import (
    DERIVED_FILL,
    MISSING_REASON_SHEET,
    SECTION_FILL_BY_STATEMENT,
    build_export_sheet_title,
    export_template_workbook,
    is_section_label,
)
from cninfo_pipeline.template_registry import TemplateSpec, discover_templates, resolve_template


def find_row_indices(sheet, label: str) -> list[int]:
    matches = [row for row in range(1, sheet.max_row + 1) if sheet.cell(row, 1).value == label]
    if not matches:
        raise AssertionError(f"missing label: {label}")
    return matches


def build_balance_record(date: str, **fields: float) -> dict:
    return {
        "ENDDATE": date,
        "F001D": date,
        "F002V": "071001",
        "F003V": "合并本期",
        "F005V": "定期报告",
        **fields,
    }


def build_statement_record(date: str, **fields: float) -> dict:
    return {
        "ENDDATE": date,
        "F001D": date,
        "F002V": "071001",
        "F003V": "合并本期",
        "F005V": "定期报告",
        **fields,
    }


def find_row_index(sheet, label: str) -> int:
    for row in range(1, sheet.max_row + 1):
        if sheet.cell(row, 1).value == label:
            return row
    raise AssertionError(f"未找到标签：{label}")


def find_column_index(sheet, label: str, *, row: int = 1) -> int:
    for column in range(1, sheet.max_column + 1):
        if sheet.cell(row, column).value == label:
            return column
    raise AssertionError(f"未找到列：{label}")


def iso_date(value: object) -> str:
    if hasattr(value, "date"):
        return value.date().isoformat()
    if hasattr(value, "isoformat"):
        return value.isoformat()
    raise AssertionError(f"不是日期值：{value!r}")


class ServiceTests(unittest.TestCase):
    def test_filter_annual_merged_records(self) -> None:
        records = [
            {"ENDDATE": "2024-12-31", "F003V": "合并本期", "F010N": 1},
            {"ENDDATE": "2024-09-30", "F003V": "合并本期", "F010N": 2},
            {"ENDDATE": "2024-12-31", "F003V": "母公司本期", "F010N": 3},
            {"ENDDATE": "2023-12-31", "F003V": "合并本期", "F010N": 4},
        ]
        filtered = filter_annual_merged_records(records)
        self.assertEqual([item["ENDDATE"] for item in filtered], ["2024-12-31", "2023-12-31"])

    def test_build_statement_matrix(self) -> None:
        records = [
            {
                "ENDDATE": "2024-12-31",
                "F006N": 100.0,
                "F008N": 10.0,
                "F009N": 20.0,
                "F010N": 30.0,
                "F011N": 40.0,
                "F015N": 50.0,
                "F018N": 60.0,
                "F019N": 70.0,
                "F025N": 80.0,
                "F026N": 90.0,
                "F033N": 5.0,
                "F037N": 1000.0,
                "F038N": 1070.0,
                "F041N": 11.0,
                "F042N": 22.0,
                "F048N": 44.0,
                "F047N": 4.0,
                "F052N": 300.0,
                "F060N": 200.0,
                "F061N": 500.0,
                "F062N": 200.0,
                "F063N": 100.0,
                "F064N": 50.0,
                "F065N": 150.0,
                "F067N": 20.0,
                "F070N": 570.0,
                "F073N": 550.0,
                "F074N": 7.0,
                "F121N": 9.0,
                "F122N": 8.0,
            }
        ]

        matrix = build_statement_matrix(records)
        rows = {row[0]: row[1:] for row in matrix}

        self.assertEqual(rows["报表日期"], [20241231])
        self.assertEqual(rows["单位"], ["元"])
        self.assertEqual(rows["货币资金"], [100.0])
        self.assertEqual(rows["应收票据及应收账款8+9"], [30.0])
        self.assertEqual(rows["在建工程(合计)32+33"], [90.0])
        self.assertEqual(rows["固定资产及清理(合计)35+36"], [80.0])
        self.assertEqual(rows["应付票据及应付账款53+54"], [33.0])
        self.assertEqual(rows["其他应付款"], [40.0])
        self.assertAlmostEqual(rows["实际资产负债率"][0], 500.0 / 1070.0)
        self.assertAlmostEqual(rows["FA&CIP占比资产总额"][0], 170.0 / 1070.0)
        self.assertAlmostEqual(rows["商誉占比归属股东权益"][0], 5.0 / 550.0)
        self.assertEqual(rows["FA&CIP净额"], [170.0])
        self.assertEqual(rows["FA&CIP净额-合计1"], [80.0])
        self.assertEqual(rows["FA&CIP净额-合计2"], [170.0])
        self.assertAlmostEqual(rows["权益乘数=资产总额/所有者权益"][0], 1070.0 / 570.0)
        self.assertAlmostEqual(rows["产权比率=负债总额/归属股东权益"][0], 500.0 / 550.0)

    def test_build_statement_matrix_supports_unit_scaling(self) -> None:
        records = [
            {
                "ENDDATE": "2024-12-31",
                "F006N": 2_500_000.0,
                "F025N": 6_000_000.0,
                "F026N": 4_000_000.0,
                "F038N": 20_000_000.0,
                "F061N": 5_000_000.0,
                "F070N": 15_000_000.0,
                "F073N": 10_000_000.0,
            }
        ]

        matrix = build_statement_matrix(records, unit_label="万元")
        rows = {row[0]: row[1:] for row in matrix}

        self.assertEqual(rows["单位"], ["万元"])
        self.assertEqual(rows["货币资金"], [250.0])
        self.assertEqual(rows["固定资产及清理(合计)35+36"], [600.0])
        self.assertEqual(rows["FA&CIP净额"], [1000.0])
        self.assertEqual(rows["资产总计"], [2000.0])
        self.assertAlmostEqual(rows["实际资产负债率"][0], 0.25)
        self.assertAlmostEqual(rows["权益乘数=资产总额/所有者权益"][0], 20_000_000.0 / 15_000_000.0)

    def test_export_statement_workbook(self) -> None:
        company = CompanyRecord(seccode="600900", secname="长江电力", orgname="中国长江电力股份有限公司")
        matrix = [
            ["报表日期", 20241231, 20231231],
            ["单位", "元", "元"],
            ["流动资产", None, None],
            ["货币资金", 100.0, 90.0],
        ]

        with tempfile.TemporaryDirectory() as tmpdir:
            workbook_path = export_statement_workbook(company, matrix, Path(tmpdir))
            self.assertTrue(workbook_path.exists())

            workbook = load_workbook(workbook_path, data_only=True)
            sheet = workbook.active
            self.assertEqual(sheet["A1"].value, "报表日期")
            self.assertEqual(sheet["B1"].value, 20241231)
            self.assertEqual(sheet["B2"].value, "元")
            self.assertEqual(sheet["A4"].value, "货币资金")
            self.assertEqual(sheet["C4"].value, 90.0)
            self.assertEqual(sheet["A1"].font.name, "等线")

    def test_prepare_cache_dir_uses_hidden_internal_dir(self) -> None:
        temp_dir = Path("test_artifacts")
        shutil.rmtree(temp_dir, ignore_errors=True)
        temp_dir.mkdir(parents=True, exist_ok=True)
        try:
            legacy_cache = temp_dir / ".cache"
            legacy_cache.mkdir(parents=True, exist_ok=True)
            (legacy_cache / "companies.json").write_text("{}", encoding="utf-8")

            cache_dir = prepare_cache_dir(temp_dir)

            self.assertEqual(cache_dir.name, ".cninfo_internal")
            self.assertTrue(cache_dir.exists())
            self.assertFalse(legacy_cache.exists())
            self.assertTrue((cache_dir / "companies.json").exists())
        finally:
            shutil.rmtree(temp_dir, ignore_errors=True)

    def test_prepare_cache_dir_defaults_to_app_cache_dir(self) -> None:
        temp_dir = Path("test_artifacts")
        shutil.rmtree(temp_dir, ignore_errors=True)
        try:
            expected = temp_dir / "app-cache"
            with patch("cninfo_pipeline.service.resolve_default_cache_dir", return_value=expected):
                cache_dir = prepare_cache_dir()

            self.assertEqual(cache_dir, expected)
            self.assertTrue(cache_dir.exists())
        finally:
            shutil.rmtree(temp_dir, ignore_errors=True)

    def test_discover_templates_finds_bank_and_company(self) -> None:
        templates = {template.template_id: template for template in discover_templates()}

        self.assertIn("银行财务报表模版", templates)
        self.assertIn("公司财务报表模版", templates)
        self.assertEqual(templates["公司财务报表模版"].kind, "company")
        self.assertEqual(resolve_template("工商银行财务报表模版").template_id, "银行财务报表模版")
        self.assertEqual(resolve_template("长江电力年度报告财务报表模版").template_id, "公司财务报表模版")

    def test_is_section_label_does_not_treat_parenthesized_items_as_titles(self) -> None:
        self.assertFalse(is_section_label("（3）现金流量套期储备", None))
        self.assertFalse(is_section_label("(2)其他权益工具投资公允价值变动", None))
        self.assertTrue(is_section_label("一、经营活动产生的现金流量：", None))
        self.assertTrue(is_section_label("资产：", None))

    def test_pipeline_run_uses_selected_template_export(self) -> None:
        company = CompanyRecord(seccode="601138", secname="工业富联", orgname="富士康工业互联网股份有限公司")
        balance_records = [build_balance_record("2024-12-31"), build_balance_record("2023-12-31")]
        income_records = [build_statement_record("2024-12-31"), build_statement_record("2023-12-31")]
        cash_flow_records = [build_statement_record("2024-12-31"), build_statement_record("2023-12-31")]
        client = MagicMock()
        client.search_company.return_value = company
        client.fetch_bank_balance_sheet.return_value = balance_records
        client.fetch_bank_income_statement.return_value = income_records
        client.fetch_bank_cash_flow_statement.return_value = cash_flow_records

        pipeline = AnnualReportPipeline(client=client)
        fake_output = Path("out") / "工业富联财务报表2024YE.xlsx"
        fake_cache_dir = Path("cache-dir")
        fake_output_dir = Path("exports")
        template_id = "银行财务报表模版"
        fake_official_provider = object()

        with (
            patch("cninfo_pipeline.service.prepare_cache_dir", return_value=fake_cache_dir),
            patch("cninfo_pipeline.service.resolve_default_output_dir", return_value=fake_output_dir),
            patch("cninfo_pipeline.service.OfficialAnnualReportSource", return_value=fake_official_provider),
            patch("cninfo_pipeline.template_export.export_template_workbook", return_value=fake_output) as exporter,
        ):
            result = pipeline.run(company_query="工业富联", unit_label="万元", template_id=template_id)

        client.set_cache_dir.assert_called_once_with(fake_cache_dir)
        client.search_company.assert_called_once_with("工业富联")
        client.fetch_bank_balance_sheet.assert_called_once_with("601138")
        client.fetch_bank_income_statement.assert_called_once_with("601138")
        client.fetch_bank_cash_flow_statement.assert_called_once_with("601138")
        client.fetch_balance_sheet.assert_not_called()
        client.fetch_income_statement.assert_not_called()
        client.fetch_cash_flow_statement.assert_not_called()
        exporter.assert_called_once_with(
            company=company,
            balance_records=balance_records,
            income_records=income_records,
            cash_flow_records=cash_flow_records,
            output_dir=fake_output_dir,
            unit_label="万元",
            template_id=template_id,
            official_provider=fake_official_provider,
        )
        self.assertEqual(result.output_path, fake_output)
        self.assertEqual(result.total_records, 2)
        self.assertEqual(result.annual_records, 2)
        self.assertEqual(result.unit_label, "万元")
        self.assertEqual(result.template_id, template_id)
        self.assertEqual(result.template_name, resolve_template(template_id).display_name)

    def test_export_template_workbook_company_template_uses_template_sheets(self) -> None:
        company = CompanyRecord(seccode="600900", secname="长江电力", orgname="中国长江电力股份有限公司")
        template = resolve_template("公司财务报表模版")
        balance_records = [
            build_balance_record("2024-12-31", F006N=10_000_000, F038N=28_000_000, F061N=15_000_000),
            build_balance_record("2023-12-31", F006N=8_000_000, F038N=24_000_000, F061N=12_000_000),
        ]
        income_records = [
            build_statement_record("2024-12-31", F006N=30_000_000, F035N=30_000_000, F018N=10_000_000, F051N=120_000),
            build_statement_record("2023-12-31", F006N=25_000_000, F035N=25_000_000, F018N=8_000_000, F051N=110_000),
        ]
        cash_flow_records = [
            build_statement_record("2024-12-31", F015N=5_000_000, F041N=9_000_000),
            build_statement_record("2023-12-31", F015N=4_000_000, F041N=8_000_000),
        ]
        official_provider = MagicMock()
        official_provider.get_statement_overrides.side_effect = lambda *_args, **kwargs: {
            "其中：利息费用": 500_000 if kwargs["period_end"] == "2024-12-31" else 400_000,
            "利息收入": 200_000 if kwargs["period_end"] == "2024-12-31" else 100_000,
            "加：其他收益": 120_000 if kwargs["period_end"] == "2024-12-31" else 110_000,
            "信用减值损失": -50_000 if kwargs["period_end"] == "2024-12-31" else -40_000,
            "资产减值损失": -30_000 if kwargs["period_end"] == "2024-12-31" else -20_000,
        }

        with tempfile.TemporaryDirectory() as tmpdir:
            workbook_path = export_template_workbook(
                company=company,
                balance_records=balance_records,
                income_records=income_records,
                cash_flow_records=cash_flow_records,
                output_dir=tmpdir,
                unit_label="万元",
                template_id=template.template_id,
                official_provider=official_provider,
            )

            workbook = load_workbook(workbook_path, data_only=False)
            template_workbook = load_workbook(template.path, data_only=False)
            self.assertEqual(
                workbook.sheetnames[:-1],
                [
                    build_export_sheet_title(company.secname, "资产负债表"),
                    build_export_sheet_title(company.secname, "利润表"),
                    build_export_sheet_title(company.secname, "现金流量表"),
                ],
            )
            self.assertEqual(workbook.sheetnames[-1], build_export_sheet_title(company.secname, MISSING_REASON_SHEET))

            balance_sheet = workbook[build_export_sheet_title(company.secname, "资产负债表")]
            balance_note_col = find_column_index(balance_sheet, "注释")
            self.assertEqual(balance_sheet["A1"].value, "报表日期")
            self.assertEqual(balance_sheet["A2"].value, "单位")
            self.assertEqual(balance_sheet.cell(1, balance_note_col).value, "注释")
            self.assertEqual(balance_sheet.cell(2, balance_note_col).value, "来源")
            self.assertEqual(balance_sheet["B2"].value, "万元")
            self.assertEqual(iso_date(balance_sheet["B1"].value), "2024-12-31")
            self.assertEqual(iso_date(balance_sheet["C1"].value), "2023-12-31")
            self.assertGreater(balance_sheet.column_dimensions["A"].width, 10)
            self.assertTrue(balance_sheet["A3"].alignment.wrap_text)
            self.assertEqual(balance_sheet["A1"].font.name, template_workbook[template_workbook.sheetnames[0]]["A1"].font.name)
            self.assertEqual(balance_sheet["A1"].font.sz, template_workbook[template_workbook.sheetnames[0]]["A1"].font.sz)
            self.assertEqual(balance_sheet["B1"].font.sz, template_workbook[template_workbook.sheetnames[0]]["A1"].font.sz)
            self.assertEqual(balance_sheet["A1"].border.left.style, "medium")
            self.assertEqual(balance_sheet.cell(1, balance_note_col).border.left.style, "medium")
            cash_row = find_row_index(balance_sheet, "货币资金")
            self.assertEqual(balance_sheet.cell(cash_row, 2).value, 1000)
            self.assertEqual(balance_sheet.cell(cash_row, 3).value, 800)

            income_sheet = workbook[build_export_sheet_title(company.secname, "利润表")]
            income_note_col = find_column_index(income_sheet, "注释")
            self.assertEqual(income_sheet["A1"].value, "报表日期")
            self.assertEqual(income_sheet.cell(1, income_note_col).value, "注释")
            self.assertEqual(iso_date(income_sheet["B1"].value), "2024-12-31")
            self.assertEqual(iso_date(income_sheet["C1"].value), "2023-12-31")
            self.assertEqual(income_sheet["A3"].value, "一、营业总收入")
            extra_gain_row = find_row_index(income_sheet, "加：其他收益")
            self.assertEqual(income_sheet.cell(extra_gain_row, 2).value, 12)
            interest_row = find_row_index(income_sheet, "其中：利息费用")
            self.assertEqual(income_sheet.cell(interest_row, income_note_col).value, "PDF年报")
            self.assertEqual(income_sheet.cell(extra_gain_row, income_note_col).value, "PDF年报")
            income_labels = {income_sheet.cell(row, 1).value for row in range(1, income_sheet.max_row + 1)}
            self.assertNotIn("营业收入", income_labels)

            cash_sheet = workbook[build_export_sheet_title(company.secname, "现金流量表")]
            cash_note_col = find_column_index(cash_sheet, "注释")
            self.assertEqual(cash_sheet["A1"].value, "报表日期")
            self.assertEqual(cash_sheet.cell(1, cash_note_col).value, "注释")
            self.assertEqual(iso_date(cash_sheet["B1"].value), "2024-12-31")
            self.assertEqual(iso_date(cash_sheet["C1"].value), "2023-12-31")

            template_workbook.close()
            workbook.close()

    def test_export_template_workbook_expands_to_all_annual_periods(self) -> None:
        company = CompanyRecord(seccode="600900", secname="长江电力", orgname="中国长江电力股份有限公司")
        template = resolve_template("公司财务报表模版")
        balance_records = [
            build_balance_record("2024-12-31", F006N=10_000_000, F038N=28_000_000, F061N=15_000_000),
            build_balance_record("2023-12-31", F006N=8_000_000, F038N=24_000_000, F061N=12_000_000),
            build_balance_record("2022-12-31", F006N=6_000_000, F038N=20_000_000, F061N=10_000_000),
        ]
        income_records = [
            build_statement_record("2024-12-31", F006N=30_000_000, F035N=30_000_000, F018N=10_000_000),
            build_statement_record("2023-12-31", F006N=25_000_000, F035N=25_000_000, F018N=8_000_000),
            build_statement_record("2022-12-31", F006N=20_000_000, F035N=20_000_000, F018N=6_000_000),
        ]
        cash_flow_records = [
            build_statement_record("2024-12-31", F015N=5_000_000, F041N=9_000_000),
            build_statement_record("2023-12-31", F015N=4_000_000, F041N=8_000_000),
            build_statement_record("2022-12-31", F015N=3_000_000, F041N=7_000_000),
        ]

        with tempfile.TemporaryDirectory() as tmpdir:
            workbook_path = export_template_workbook(
                company=company,
                balance_records=balance_records,
                income_records=income_records,
                cash_flow_records=cash_flow_records,
                output_dir=tmpdir,
                unit_label="万元",
                template_id=template.template_id,
            )

            workbook = load_workbook(workbook_path, data_only=False)
            balance_sheet = workbook[workbook.sheetnames[0]]
            note_col = find_column_index(balance_sheet, "注释")
            self.assertEqual(balance_sheet.cell(1, note_col).value, "注释")
            self.assertEqual(iso_date(balance_sheet["B1"].value), "2024-12-31")
            self.assertEqual(iso_date(balance_sheet["C1"].value), "2023-12-31")
            self.assertEqual(iso_date(balance_sheet["D1"].value), "2022-12-31")
            cash_row = find_row_index(balance_sheet, "货币资金")
            self.assertEqual(balance_sheet.cell(cash_row, 4).value, 600)
            workbook.close()

    def test_extract_company_income_values_from_text_uses_official_pdf_rows(self) -> None:
        text = """
中国石油天然气股份有限公司
2024 年度合并及公司利润表
(除特别注明外，金额单位为人民币百万元)
财务费用  47  (12,552)  (18,091)
其中：利息费用    20,731  24,063
利息收入    8,799  8,288
加：其他收益  48  20,122  21,704
投资收益  49  11,934  9,554
信用减值损失  51  (742)  (35)
资产减值损失  52  (14,278)  (28,956)
资产处置收益  53  613  498
"""
        values = extract_company_income_values_from_text(text)
        self.assertEqual(values["利息费用"], 20_731_000_000)
        self.assertEqual(values["利息收入"], 8_799_000_000)
        self.assertEqual(values["其他收益"], 20_122_000_000)
        self.assertEqual(values["信用减值损失"], -742_000_000)
        self.assertEqual(values["资产减值损失"], -14_278_000_000)
        self.assertEqual(values["资产处置收益"], 613_000_000)

    def test_extract_company_income_values_from_text_avoids_topline_interest_income(self) -> None:
        text = """
贵州茅台酒股份有限公司 2024 年年度报告
合并利润表
2024 年1—12 月
单位：元  币种：人民币
一、营业总收入  174,144,069,958.25 150,560,330,316.45
其中：营业收入 40 170,899,152,276.34 147,693,604,994.14
利息收入 41 3,244,917,681.91 2,866,725,322.31
二、营业总成本  54,523,971,452.57 46,960,889,468.54
财务费用 46 -1,470,219,863.34 -1,789,503,701.48
其中：利息费用  14,474,584.09 12,624,628.35
利息收入  1,476,991,223.18 1,942,301,920.98
加：其他收益 47 21,229,466.81 34,644,873.86
信用减值损失（损失以“-”号填列） 50 -23,248,436.03 37,871,293.26
资产减值损失（损失以“-”号填列）
资产处置收益（损失以“－”号填列） 51 388,852.05 -479,736.97
"""
        values = extract_company_income_values_from_text(text)
        self.assertEqual(values["利息费用"], 14_474_584.09)
        self.assertEqual(values["利息收入"], 1_476_991_223.18)
        self.assertEqual(values["其他收益"], 21_229_466.81)
        self.assertEqual(values["信用减值损失"], -23_248_436.03)
        self.assertNotIn("资产减值损失", values)

    def test_extract_bank_balance_values_from_text_uses_official_pdf_rows(self) -> None:
        text = """
人民币百万元，百分比除外
客户存款 37,311,778 34,836,973 33,521,174
同业及其他金融机构存放款项 4,568,696 4,020,537 2,841,385
拆入资金 534,551 570,428 528,473
客户贷款及垫款净额(1) 29,712,359 55.6 27,613,781 56.6
现金及存放中央银行款项 3,674,558 6.9 3,322,911 6.8
买入返售款项 530,737 1.0 1,210,217 2.5
"""
        values = extract_bank_balance_values_from_text(text)
        self.assertEqual(values["客户存款(吸收存款)"], 37_311_778_000_000)
        self.assertEqual(values["其中:同业存放款项"], 4_568_696_000_000)
        self.assertEqual(values["拆入资金"], 534_551_000_000)
        self.assertEqual(values["发放贷款及垫款"], 29_712_359_000_000)
        self.assertEqual(values["现金及存放中央银行款项"], 3_674_558_000_000)
        self.assertEqual(values["买入返售金融资产"], 530_737_000_000)

    def test_extract_statement_values_from_text_matches_company_balance_labels(self) -> None:
        text = """
合并资产负债表
单位：万元
货币资金 123,456 100,000
应收票据 4,321 3,210
"""
        values = extract_statement_values_from_text(
            text,
            template_kind="company",
            statement_type="balance",
            requested_labels=["货币资金", "应收票据"],
        )
        self.assertEqual(values["货币资金"], 1_234_560_000)
        self.assertEqual(values["应收票据"], 43_210_000)

    def test_extract_statement_values_from_text_matches_bank_income_labels(self) -> None:
        text = """
合并利润表
单位：百万元
利息收入 1,234 1,111
手续费及佣金收入 222 200
"""
        values = extract_statement_values_from_text(
            text,
            template_kind="bank",
            statement_type="income",
            requested_labels=["其中：利息收入", "其中:手续费及佣金收入"],
        )
        self.assertEqual(values["其中：利息收入"], 1_234_000_000)
        self.assertEqual(values["其中:手续费及佣金收入"], 222_000_000)

    def test_extract_statement_values_from_text_prefers_exact_label_match(self) -> None:
        text = (
            "\u5408\u5e76\u8d44\u4ea7\u8d1f\u503a\u8868\n"
            "\u5355\u4f4d\uff1a\u4e07\u5143\n"
            "\u6d41\u52a8\u8d1f\u503a\u5408\u8ba1 100 90\n"
            "\u8d1f\u503a\u5408\u8ba1 200 180\n"
        )
        values = extract_statement_values_from_text(
            text,
            template_kind="company",
            statement_type="balance",
            requested_labels=["\u8d1f\u503a\u5408\u8ba1"],
        )
        self.assertEqual(values["\u8d1f\u503a\u5408\u8ba1"], 2_000_000)

    def test_export_template_workbook_bank_template_inserts_year_headers(self) -> None:
        company = CompanyRecord(seccode="601398", secname="工商银行", orgname="中国工商银行股份有限公司")
        balance_records = [
            build_balance_record(
                "2025-12-31",
                F034N=3_000_000,
                F048N=2_100_000,
                F049N=3_600_000,
                F052N=1_200_000,
                F054N=1_300_000,
                F082N=16_000_000,
                F083N=500_000,
                F084N=600_000,
                F086N=700_000,
                F087N=900_000,
                F088N=800_000,
                F089N=90_000,
                F092N=5_000_000,
                F093N=21_000_000,
                F094N=4_910_000,
                F097N=1_400_000,
                F098N=500_000,
                F099N=900_000,
                F100N=110_000,
            ),
            build_balance_record(
                "2024-12-31",
                F034N=2_500_000,
                F048N=1_900_000,
                F049N=3_300_000,
                F052N=1_000_000,
                F054N=1_100_000,
                F082N=14_000_000,
                F083N=450_000,
                F084N=550_000,
                F086N=650_000,
                F087N=850_000,
                F088N=750_000,
                F089N=80_000,
                F092N=4_500_000,
                F093N=18_500_000,
                F094N=4_420_000,
                F097N=1_200_000,
                F098N=400_000,
                F099N=800_000,
                F100N=90_000,
            ),
        ]
        income_records = [
            build_statement_record(
                "2025-12-31",
                F006N=900_000,
                F007N=500_000,
                F008N=800_000,
                F009N=300_000,
                F048N=300_000,
            ),
            build_statement_record(
                "2024-12-31",
                F006N=800_000,
                F007N=400_000,
                F008N=700_000,
                F009N=300_000,
                F048N=250_000,
            ),
        ]
        cash_flow_records = [
            build_statement_record(
                "2025-12-31",
                F010N=120_000,
                F019N=30_000,
                F021N=130_000,
                F027N=40_000,
                F028N=20_000,
                F031N=100_000,
                F033N=60_000,
                F034N=70_000,
                F039N=50_000,
                F045N=110_000,
                F047N=70_000,
                F054N=60_000,
                F057N=500_000,
                F058N=400_000,
            ),
            build_statement_record(
                "2024-12-31",
                F010N=110_000,
                F019N=25_000,
                F021N=120_000,
                F027N=35_000,
                F028N=15_000,
                F031N=70_000,
                F033N=55_000,
                F034N=65_000,
                F039N=45_000,
                F045N=100_000,
                F047N=65_000,
                F054N=55_000,
                F057N=450_000,
                F058N=350_000,
            ),
        ]

        with tempfile.TemporaryDirectory() as tmpdir:
            template_path = Path(tmpdir) / "bank_template.xlsx"
            template_workbook = Workbook()

            balance_sheet = template_workbook.active
            balance_sheet.title = "测试银行资产负债表"
            for row, label in enumerate(
                [
                    "项目资产",
                    "长期股权投资",
                    "向中央银行借款",
                    "同业和其他金融机构存放款项",
                    "拆入资金",
                    "衍生金融负债",
                    "其他权益工具",
                    "其中：优先股",
                    "永续债",
                    "资本公积",
                    "其他综合收益",
                    "盈余公积",
                    "一般准备",
                    "未分配利润",
                    "少数股东权益",
                    "负债合计",
                ],
                start=1,
            ):
                balance_sheet.cell(row, 1, label)

            income_sheet = template_workbook.create_sheet("测试银行利润表")
            for row, label in enumerate(["利息净收入", "营业收入小计", "净利润"], start=1):
                income_sheet.cell(row, 1, label)

            cash_sheet = template_workbook.create_sheet("测试银行现金流量表")
            for row, label in enumerate(
                [
                    "一、经营活动产生的现金流量",
                    "收取的利息、手续费及佣金的现金",
                    "贷款和垫款净增加额",
                    "支付给职工以及为职工支付的现金",
                    "支付的各项税费",
                    "收到其他与经营活动有关的现金",
                    "收回投资收到的现金",
                    "取得投资收益收到的现金",
                    "购建固定资产和其他资产所支付的现金",
                    "发行债券收到的现金",
                    "收到其他与筹资活动有关的现金",
                    "四、汇率变动对现金及现金等价物的影响额",
                    "经营活动产生的现金流量净额",
                    "加：年初现金及现金等价物余额",
                    "年末现金及现金等价物余额",
                ],
                start=1,
            ):
                cash_sheet.cell(row, 1, label)

            template_workbook.save(template_path)
            template_workbook.close()

            template = TemplateSpec(
                template_id="bank-test",
                display_name="银行财务 - 测试模板",
                kind="bank",
                path=template_path,
            )
            official_provider = MagicMock()
            official_provider.get_statement_overrides.return_value = {}

            with patch("cninfo_pipeline.template_export.resolve_template", return_value=template):
                workbook_path = export_template_workbook(
                    company=company,
                    balance_records=balance_records,
                    income_records=income_records,
                    cash_flow_records=cash_flow_records,
                    output_dir=tmpdir,
                    unit_label="万元",
                    template_id=template.template_id,
                    official_provider=official_provider,
                )

            workbook = load_workbook(workbook_path, data_only=False)
            balance_sheet = workbook[build_export_sheet_title(company.secname, "资产负债表")]
            income_sheet = workbook[build_export_sheet_title(company.secname, "利润表")]
            cash_sheet = workbook[build_export_sheet_title(company.secname, "现金流量表")]
            reason_sheet = workbook[build_export_sheet_title(company.secname, MISSING_REASON_SHEET)]

            self.assertEqual(balance_sheet["A1"].value, "项目（单位：万元）")
            balance_note_col = find_column_index(balance_sheet, "注释")
            self.assertEqual(balance_sheet.cell(1, balance_note_col).value, "注释")
            self.assertEqual(iso_date(balance_sheet["B1"].value), "2025-12-31")
            self.assertEqual(iso_date(balance_sheet["C1"].value), "2024-12-31")
            self.assertEqual(balance_sheet["A2"].value, "项目资产")
            self.assertGreater(balance_sheet.column_dimensions["A"].width, 10)
            self.assertTrue(balance_sheet["A2"].alignment.wrap_text)
            self.assertTrue(balance_sheet["A2"].fill.fgColor.rgb.endswith(SECTION_FILL_BY_STATEMENT["balance"].fgColor.rgb[-6:]))
            self.assertTrue(balance_sheet["B2"].fill.fgColor.rgb.endswith(SECTION_FILL_BY_STATEMENT["balance"].fgColor.rgb[-6:]))
            self.assertIsNone(balance_sheet["A2"].comment)
            self.assertEqual(balance_sheet["A1"].font.sz, 11)
            self.assertEqual(balance_sheet["B1"].font.sz, 11)
            self.assertEqual(balance_sheet["A1"].border.left.style, "medium")
            self.assertEqual(balance_sheet.cell(1, balance_note_col).border.left.style, "medium")
            long_term_row = find_row_index(balance_sheet, "长期股权投资")
            self.assertEqual(balance_sheet.cell(long_term_row, 2).value, 300)
            cb_borrow_row = find_row_index(balance_sheet, "向中央银行借款")
            self.assertEqual(balance_sheet.cell(cb_borrow_row, 2).value, 210)
            peer_deposit_row = find_row_index(balance_sheet, "同业和其他金融机构存放款项")
            self.assertEqual(balance_sheet.cell(peer_deposit_row, 2).value, 360)
            split_borrow_row = find_row_index(balance_sheet, "拆入资金")
            self.assertEqual(balance_sheet.cell(split_borrow_row, 2).value, 120)
            derivative_liab_row = find_row_index(balance_sheet, "衍生金融负债")
            self.assertEqual(balance_sheet.cell(derivative_liab_row, 2).value, 130)
            equity_tool_row = find_row_index(balance_sheet, "其他权益工具")
            self.assertEqual(balance_sheet.cell(equity_tool_row, 2).value, 140)
            preferred_row = find_row_index(balance_sheet, "其中：优先股")
            self.assertEqual(balance_sheet.cell(preferred_row, 2).value, 50)
            perpetual_row = find_row_index(balance_sheet, "永续债")
            self.assertEqual(balance_sheet.cell(perpetual_row, 2).value, 90)
            oci_row = find_row_index(balance_sheet, "其他综合收益")
            self.assertEqual(balance_sheet.cell(oci_row, 2).value, 11)
            minority_row = find_row_index(balance_sheet, "少数股东权益")
            self.assertEqual(balance_sheet.cell(minority_row, 2).value, 9)

            self.assertEqual(income_sheet["A2"].value, "利息净收入")
            income_note_col = find_column_index(income_sheet, "注释")
            self.assertEqual(income_sheet.cell(1, income_note_col).value, "注释")
            self.assertEqual(iso_date(income_sheet["B1"].value), "2025-12-31")
            interest_row = find_row_index(income_sheet, "利息净收入")
            self.assertEqual(income_sheet.cell(interest_row, 2).value, 50)
            revenue_row = find_row_index(income_sheet, "营业收入小计")
            self.assertEqual(income_sheet.cell(revenue_row, 2).value, 90)
            profit_row = find_row_index(income_sheet, "净利润")
            self.assertEqual(income_sheet.cell(profit_row, 2).value, 30)

            self.assertEqual(cash_sheet["A2"].value, "一、经营活动产生的现金流量")
            cash_note_col = find_column_index(cash_sheet, "注释")
            self.assertEqual(cash_sheet.cell(1, cash_note_col).value, "注释")
            self.assertEqual(iso_date(cash_sheet["B1"].value), "2025-12-31")
            self.assertTrue(cash_sheet["A2"].fill.fgColor.rgb.endswith(SECTION_FILL_BY_STATEMENT["cash"].fgColor.rgb[-6:]))
            self.assertTrue(cash_sheet["B2"].fill.fgColor.rgb.endswith(SECTION_FILL_BY_STATEMENT["cash"].fgColor.rgb[-6:]))
            collected_fee_row = find_row_index(cash_sheet, "收取的利息、手续费及佣金的现金")
            self.assertEqual(cash_sheet.cell(collected_fee_row, 2).value, 12)
            loans_row = find_row_index(cash_sheet, "贷款和垫款净增加额")
            self.assertEqual(cash_sheet.cell(loans_row, 2).value, 13)
            payroll_row = find_row_index(cash_sheet, "支付给职工以及为职工支付的现金")
            self.assertEqual(cash_sheet.cell(payroll_row, 2).value, 4)
            tax_row = find_row_index(cash_sheet, "支付的各项税费")
            self.assertEqual(cash_sheet.cell(tax_row, 2).value, 2)
            other_operating_row = find_row_index(cash_sheet, "收到其他与经营活动有关的现金")
            self.assertEqual(cash_sheet.cell(other_operating_row, 2).value, 3)
            other_invest_row = find_row_index(cash_sheet, "收回投资收到的现金")
            self.assertEqual(cash_sheet.cell(other_invest_row, 2).value, 6)
            invest_gain_row = find_row_index(cash_sheet, "取得投资收益收到的现金")
            self.assertEqual(cash_sheet.cell(invest_gain_row, 2).value, 7)
            capex_row = find_row_index(cash_sheet, "购建固定资产和其他资产所支付的现金")
            self.assertEqual(cash_sheet.cell(capex_row, 2).value, 5)
            debt_issue_row = find_row_index(cash_sheet, "发行债券收到的现金")
            self.assertEqual(cash_sheet.cell(debt_issue_row, 2).value, 11)
            other_finance_row = find_row_index(cash_sheet, "收到其他与筹资活动有关的现金")
            self.assertEqual(cash_sheet.cell(other_finance_row, 2).value, 7)
            fx_row = find_row_index(cash_sheet, "四、汇率变动对现金及现金等价物的影响额")
            self.assertEqual(cash_sheet.cell(fx_row, 2).value, 6)
            net_cash_row = find_row_index(cash_sheet, "经营活动产生的现金流量净额")
            self.assertEqual(cash_sheet.cell(net_cash_row, 2).value, 10)
            opening_cash_row = find_row_index(cash_sheet, "加：年初现金及现金等价物余额")
            self.assertEqual(cash_sheet.cell(opening_cash_row, 2).value, 50)
            closing_cash_row = find_row_index(cash_sheet, "年末现金及现金等价物余额")
            self.assertEqual(cash_sheet.cell(closing_cash_row, 2).value, 40)
            self.assertEqual(reason_sheet["A1"].value, "工作表")
            reason_labels = {reason_sheet.cell(row, 3).value for row in range(2, reason_sheet.max_row + 1)}
            self.assertIn("项目资产", reason_labels)
            self.assertIn("一、经营活动产生的现金流量", reason_labels)

            workbook.close()

    def test_export_template_workbook_company_balance_prefers_api_over_official_rows(self) -> None:
        company = CompanyRecord(seccode="600900", secname="长江电力", orgname="中国长江电力股份有限公司")
        template = resolve_template("公司")
        balance_records = [
            build_balance_record("2024-12-31", F006N=10_000_000, F038N=28_000_000, F061N=15_000_000),
            build_balance_record("2023-12-31", F006N=8_000_000, F038N=24_000_000, F061N=12_000_000),
        ]
        income_records = [
            build_statement_record("2024-12-31", F006N=30_000_000, F035N=30_000_000, F018N=10_000_000),
            build_statement_record("2023-12-31", F006N=25_000_000, F035N=25_000_000, F018N=8_000_000),
        ]
        cash_flow_records = [
            build_statement_record("2024-12-31", F015N=5_000_000, F041N=9_000_000),
            build_statement_record("2023-12-31", F015N=4_000_000, F041N=8_000_000),
        ]
        official_provider = MagicMock()

        def provider(_company, *, statement_type: str, period_end: str, **_kwargs):
            if statement_type != "balance":
                return {}
            return {"货币资金": 12_000_000 if period_end == "2024-12-31" else 9_000_000}

        official_provider.get_statement_overrides.side_effect = provider

        with tempfile.TemporaryDirectory() as tmpdir:
            workbook_path = export_template_workbook(
                company=company,
                balance_records=balance_records,
                income_records=income_records,
                cash_flow_records=cash_flow_records,
                output_dir=tmpdir,
                unit_label="万元",
                template_id=template.template_id,
                official_provider=official_provider,
            )

            workbook = load_workbook(workbook_path, data_only=False)
            balance_sheet = workbook[build_export_sheet_title(company.secname, "资产负债表")]
            note_col = find_column_index(balance_sheet, "注释")
            cash_row = find_row_index(balance_sheet, "货币资金")
            self.assertEqual(balance_sheet.cell(cash_row, 2).value, 1000)
            self.assertEqual(balance_sheet.cell(cash_row, 3).value, 800)
            self.assertEqual(balance_sheet.cell(cash_row, note_col).value, "API接口")
            workbook.close()

    def test_export_template_workbook_falls_back_to_api_when_official_income_missing(self) -> None:
        company = CompanyRecord(seccode="600900", secname="长江电力", orgname="中国长江电力股份有限公司")
        template = resolve_template("公司")
        balance_records = [
            build_balance_record("2024-12-31", F006N=10_000_000, F038N=28_000_000, F061N=15_000_000),
        ]
        income_records = [
            build_statement_record(
                "2024-12-31",
                F006N=30_000_000,
                F035N=30_000_000,
                F051N=12_000_000,
                F059N=200_000,
                F064N=-300_000,
                F065N=-500_000,
            ),
        ]
        cash_flow_records = [build_statement_record("2024-12-31", F015N=5_000_000, F041N=9_000_000)]
        official_provider = MagicMock()
        official_provider.get_statement_overrides.return_value = {}

        with tempfile.TemporaryDirectory() as tmpdir:
            workbook_path = export_template_workbook(
                company=company,
                balance_records=balance_records,
                income_records=income_records,
                cash_flow_records=cash_flow_records,
                output_dir=tmpdir,
                unit_label="万元",
                template_id=template.template_id,
                official_provider=official_provider,
            )

            workbook = load_workbook(workbook_path, data_only=False)
            income_sheet = workbook[build_export_sheet_title(company.secname, "利润表")]
            note_col = find_column_index(income_sheet, "注释")
            for label, expected in {
                "加：其他收益": 1200,
                "信用减值损失": -30,
                "资产减值损失": -50,
                "资产处置收益": 20,
            }.items():
                row = find_row_index(income_sheet, label)
                self.assertEqual(income_sheet.cell(row, 2).value, expected)
                self.assertEqual(income_sheet.cell(row, note_col).value, "API接口")
            workbook.close()

    def test_export_template_workbook_marks_pdf_api_conflict(self) -> None:
        company = CompanyRecord(seccode="600900", secname="长江电力", orgname="中国长江电力股份有限公司")
        template = resolve_template("公司")
        balance_records = [build_balance_record("2024-12-31", F006N=10_000_000, F038N=28_000_000, F061N=15_000_000)]
        income_records = [
            build_statement_record(
                "2024-12-31",
                F006N=30_000_000,
                F035N=30_000_000,
                F051N=12_000_000,
            )
        ]
        cash_flow_records = [build_statement_record("2024-12-31", F015N=5_000_000, F041N=9_000_000)]
        official_provider = MagicMock()
        official_provider.get_statement_overrides.return_value = {"加：其他收益": 13_000_000}

        with tempfile.TemporaryDirectory() as tmpdir:
            workbook_path = export_template_workbook(
                company=company,
                balance_records=balance_records,
                income_records=income_records,
                cash_flow_records=cash_flow_records,
                output_dir=tmpdir,
                unit_label="万元",
                template_id=template.template_id,
                official_provider=official_provider,
            )

            workbook = load_workbook(workbook_path, data_only=False)
            income_sheet = workbook[build_export_sheet_title(company.secname, "利润表")]
            note_col = find_column_index(income_sheet, "注释")
            row = find_row_index(income_sheet, "加：其他收益")
            self.assertEqual(income_sheet.cell(row, 2).value, 1300)
            self.assertEqual(income_sheet.cell(row, note_col).value, "PDF年报（与API不一致）")
            workbook.close()

    def test_export_template_workbook_uses_occurrence_specific_company_balance_resolvers(self) -> None:
        company = CompanyRecord(
            seccode="600900",
            secname="\u957f\u6c5f\u7535\u529b",
            orgname="\u4e2d\u56fd\u957f\u6c5f\u7535\u529b\u80a1\u4efd\u6709\u9650\u516c\u53f8",
        )
        template = resolve_template("\u516c\u53f8")
        balance_records = [
            build_balance_record(
                "2024-12-31",
                F011N=3_000_000,
                F014N=4_000_000,
                F046N=1_000_000,
                F047N=2_000_000,
                F048N=10_000_000,
                F038N=28_000_000,
                F061N=15_000_000,
            )
        ]
        with tempfile.TemporaryDirectory() as tmpdir:
            workbook_path = export_template_workbook(
                company=company,
                balance_records=balance_records,
                income_records=[],
                cash_flow_records=[],
                output_dir=tmpdir,
                unit_label="\u4e07\u5143",
                template_id=template.template_id,
            )

            workbook = load_workbook(workbook_path, data_only=False)
            balance_sheet = workbook[build_export_sheet_title(company.secname, "\u8d44\u4ea7\u8d1f\u503a\u8868")]

            receivable_rows = find_row_indices(balance_sheet, "\u5176\u4ed6\u5e94\u6536\u6b3e")
            self.assertEqual(balance_sheet.cell(receivable_rows[0], 2).value, 300)
            self.assertEqual(balance_sheet.cell(receivable_rows[1], 2).value, 400)

            payable_rows = find_row_indices(balance_sheet, "\u5176\u4ed6\u5e94\u4ed8\u6b3e")
            self.assertEqual(balance_sheet.cell(payable_rows[0], 2).value, 1000)
            self.assertEqual(balance_sheet.cell(payable_rows[1], 2).value, 700)
            workbook.close()

    def test_export_template_workbook_company_balance_hides_duplicate_parent_rows(self) -> None:
        company = CompanyRecord(seccode="600900", secname="长江电力", orgname="中国长江电力股份有限公司")
        template = resolve_template("公司")
        balance_records = [
            build_balance_record(
                "2024-12-31",
                F008N=1_000_000,
                F009N=2_000_000,
                F025N=30_000_000,
                F041N=4_000_000,
                F042N=5_000_000,
                F043N=7_000_000,
                F115N=7_000_000,
                F038N=100_000_000,
                F061N=60_000_000,
            )
        ]

        with tempfile.TemporaryDirectory() as tmpdir:
            workbook_path = export_template_workbook(
                company=company,
                balance_records=balance_records,
                income_records=[],
                cash_flow_records=[],
                output_dir=tmpdir,
                unit_label="万元",
                template_id=template.template_id,
            )

            workbook = load_workbook(workbook_path, data_only=False)
            balance_sheet = workbook[build_export_sheet_title(company.secname, "资产负债表")]
            note_col = find_column_index(balance_sheet, "注释")

            for label in ("应收票据及应收账款", "固定资产及清理", "应付票据及应付账款", "预收款项"):
                row = find_row_index(balance_sheet, label)
                self.assertIsNone(balance_sheet.cell(row, 2).value)
                self.assertIsNone(balance_sheet.cell(row, note_col).value)

            self.assertEqual(balance_sheet.cell(find_row_index(balance_sheet, "应收票据"), 2).value, 100)
            self.assertEqual(balance_sheet.cell(find_row_index(balance_sheet, "应收账款"), 2).value, 200)
            self.assertEqual(balance_sheet.cell(find_row_index(balance_sheet, "固定资产净额"), 2).value, 3000)
            self.assertEqual(balance_sheet.cell(find_row_index(balance_sheet, "应付票据"), 2).value, 400)
            self.assertEqual(balance_sheet.cell(find_row_index(balance_sheet, "应付账款"), 2).value, 500)
            self.assertEqual(balance_sheet.cell(find_row_index(balance_sheet, "合同负债"), 2).value, 700)
            workbook.close()

    def test_export_template_workbook_company_balance_uses_api_when_pdf_split_rows_conflict(self) -> None:
        company = CompanyRecord(seccode="600900", secname="长江电力", orgname="中国长江电力股份有限公司")
        template = resolve_template("公司")
        balance_records = [
            build_balance_record(
                "2024-12-31",
                F008N=1_000_000,
                F009N=2_000_000,
                F012N=300_000,
                F013N=400_000,
                F041N=500_000,
                F042N=6_000_000,
                F046N=700_000,
                F047N=800_000,
                F038N=100_000_000,
                F061N=60_000_000,
            )
        ]
        official_provider = MagicMock()
        official_provider.get_statement_overrides.return_value = {
            "应收票据": 3_000_000,
            "应收账款": 3_000_000,
            "应收利息": 500_000,
            "应收股利": 500_000,
            "应付票据": 6_500_000,
            "应付账款": 6_500_000,
            "应付利息": 1_500_000,
            "应付股利": 1_500_000,
        }

        with tempfile.TemporaryDirectory() as tmpdir:
            workbook_path = export_template_workbook(
                company=company,
                balance_records=balance_records,
                income_records=[],
                cash_flow_records=[],
                output_dir=tmpdir,
                unit_label="万元",
                template_id=template.template_id,
                official_provider=official_provider,
            )

            workbook = load_workbook(workbook_path, data_only=False)
            balance_sheet = workbook[build_export_sheet_title(company.secname, "资产负债表")]
            note_col = find_column_index(balance_sheet, "注释")

            for label, expected in {
                "应收票据": 100,
                "应收账款": 200,
                "应收利息": 30,
                "应收股利": 40,
                "应付票据": 50,
                "应付账款": 600,
                "应付利息": 70,
                "应付股利": 80,
            }.items():
                row = find_row_index(balance_sheet, label)
                self.assertEqual(balance_sheet.cell(row, 2).value, expected)
                self.assertEqual(balance_sheet.cell(row, note_col).value, "API接口")

            workbook.close()

    def test_export_template_workbook_company_balance_ignores_pdf_for_direct_rows(self) -> None:
        company = CompanyRecord(seccode="600900", secname="长江电力", orgname="中国长江电力股份有限公司")
        template = resolve_template("公司")
        balance_records = [
            build_balance_record(
                "2024-12-31",
                F010N=900_000,
                F011N=3_000_000,
                F014N=1_500_000,
                F015N=800_000,
                F023N=7_000_000,
                F038N=100_000_000,
                F042N=5_000_000,
                F047N=2_000_000,
                F048N=10_000_000,
                F061N=60_000_000,
                F115N=4_000_000,
            )
        ]
        official_provider = MagicMock()
        official_provider.get_statement_overrides.return_value = {
            "交易性金融资产": 2_000_000,
            "衍生金融资产": 2_000_000,
            "应收款项融资": 900_000,
            "预付款项": 900_000,
            "买入返售金融资产": 800_000,
            "存货": 800_000,
            "长期应收款": 7_000_000,
            "长期股权投资": 7_000_000,
            "预收款项": 4_000_000,
            "合同负债": 4_000_000,
        }

        with tempfile.TemporaryDirectory() as tmpdir:
            workbook_path = export_template_workbook(
                company=company,
                balance_records=balance_records,
                income_records=[],
                cash_flow_records=[],
                output_dir=tmpdir,
                unit_label="万元",
                template_id=template.template_id,
                official_provider=official_provider,
            )

            workbook = load_workbook(workbook_path, data_only=False)
            balance_sheet = workbook[build_export_sheet_title(company.secname, "资产负债表")]
            note_col = find_column_index(balance_sheet, "注释")

            self.assertIsNone(balance_sheet.cell(find_row_index(balance_sheet, "交易性金融资产"), 2).value)
            self.assertEqual(balance_sheet.cell(find_row_index(balance_sheet, "衍生金融资产"), 2).value, 200)
            self.assertEqual(balance_sheet.cell(find_row_index(balance_sheet, "预付款项"), 2).value, 90)
            self.assertEqual(balance_sheet.cell(find_row_index(balance_sheet, "存货"), 2).value, 80)
            self.assertEqual(balance_sheet.cell(find_row_index(balance_sheet, "长期股权投资"), 2).value, 700)
            self.assertEqual(balance_sheet.cell(find_row_index(balance_sheet, "合同负债"), 2).value, 400)
            self.assertEqual(balance_sheet.cell(find_row_index(balance_sheet, "衍生金融资产"), note_col).value, "PDF年报")
            self.assertEqual(balance_sheet.cell(find_row_index(balance_sheet, "预付款项"), note_col).value, "API接口")
            self.assertEqual(balance_sheet.cell(find_row_index(balance_sheet, "存货"), note_col).value, "API接口")

            workbook.close()

    def test_export_template_workbook_company_balance_fills_verified_pdf_only_rows(self) -> None:
        company = CompanyRecord(seccode="600900", secname="长江电力", orgname="中国长江电力股份有限公司")
        template = resolve_template("公司")
        balance_records = [
            build_balance_record(
                "2024-12-31",
                F038N=100_000_000,
                F061N=60_000_000,
            )
        ]
        official_provider = MagicMock()
        official_provider.get_statement_overrides.return_value = {
            "衍生金融资产": 2_000_000,
            "应收股利": 3_000_000,
            "合同资产": 4_000_000,
            "债权投资": 5_000_000,
            "应付票据": 6_000_000,
        }

        with tempfile.TemporaryDirectory() as tmpdir:
            workbook_path = export_template_workbook(
                company=company,
                balance_records=balance_records,
                income_records=[],
                cash_flow_records=[],
                output_dir=tmpdir,
                unit_label="万元",
                template_id=template.template_id,
                official_provider=official_provider,
            )

            workbook = load_workbook(workbook_path, data_only=False)
            balance_sheet = workbook[build_export_sheet_title(company.secname, "资产负债表")]
            note_col = find_column_index(balance_sheet, "注释")

            for label, expected in {
                "衍生金融资产": 200,
                "应收股利": 300,
                "合同资产": 400,
                "债权投资": 500,
                "应付票据": 600,
            }.items():
                row = find_row_index(balance_sheet, label)
                self.assertEqual(balance_sheet.cell(row, 2).value, expected)
                self.assertEqual(balance_sheet.cell(row, note_col).value, "PDF年报")

            workbook.close()

    def test_export_template_workbook_explains_missing_rows(self) -> None:
        company = CompanyRecord(seccode="601398", secname="工商银行", orgname="中国工商银行股份有限公司")
        balance_records = [
            build_balance_record("2025-12-31", F084N=None),
            build_balance_record("2024-12-31", F084N=None),
        ]

        with tempfile.TemporaryDirectory() as tmpdir:
            template_path = Path(tmpdir) / "bank_missing_template.xlsx"
            template_workbook = Workbook()
            sheet = template_workbook.active
            sheet.title = "测试银行资产负债表"
            for row, label in enumerate(["资产：", "未映射项目", "优先股", "资本公积"], start=1):
                sheet.cell(row, 1, label)
            template_workbook.save(template_path)
            template_workbook.close()

            template = TemplateSpec(
                template_id="bank-missing-test",
                display_name="银行财务 - 缺口测试模板",
                kind="bank",
                path=template_path,
            )

            with patch("cninfo_pipeline.template_export.resolve_template", return_value=template):
                workbook_path = export_template_workbook(
                    company=company,
                    balance_records=balance_records,
                    income_records=[],
                    cash_flow_records=[],
                    output_dir=tmpdir,
                    unit_label="万元",
                    template_id=template.template_id,
                )

            workbook = load_workbook(workbook_path, data_only=False)
            reason_sheet = workbook[build_export_sheet_title(company.secname, MISSING_REASON_SHEET)]
            reasons = {
                reason_sheet.cell(row, 3).value: reason_sheet.cell(row, 4).value
                for row in range(2, reason_sheet.max_row + 1)
            }

            self.assertEqual(reasons["资产："], "模板分组行")
            self.assertEqual(reasons["未映射项目"], "模板有此行但接口无对应字段")
            self.assertEqual(reasons["优先股"], "模板有此行但接口无对应字段")
            self.assertEqual(reasons["资本公积"], "接口字段存在但所选年度无值")
            workbook.close()


if __name__ == "__main__":
    unittest.main()
